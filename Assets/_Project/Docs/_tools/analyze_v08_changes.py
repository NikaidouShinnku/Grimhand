#!/usr/bin/env python3
"""Produce clean v0.8 change list vs repo snapshot + code gaps."""
from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
XLSX = Path(r"c:\Users\Kelthuzad\Desktop\Grimhand实际内容总览表v0.8.xlsx")
OLD = json.loads((ROOT / "Docs" / "_v08_excel_full.json").read_text(encoding="utf-8"))
OUT = ROOT / "Docs" / "_v08_implementation_checklist.md"


def s(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def load_player_cards_xlsx() -> dict[str, dict]:
    ws = openpyxl.load_workbook(XLSX, data_only=True)["卡牌"]
    out: dict[str, dict] = {}
    for r in range(2, ws.max_row + 1):
        name = s(ws.cell(r, 2).value)
        role = s(ws.cell(r, 1).value)
        if not name or name == "卡牌名称":
            continue
        if role not in ("战士", "法老", "恶魔"):
            continue
        out[name] = {
            "role": role,
            "cost": s(ws.cell(r, 3).value),
            "type": s(ws.cell(r, 4).value),
            "desc": s(ws.cell(r, 6).value),
            "rarity": s(ws.cell(r, 7).value),
            "max_upgrade": s(ws.cell(r, 8).value),
            "upgrade_per": s(ws.cell(r, 9).value),
            "upgrade_xp": s(ws.cell(r, 10).value),
            "special_effect": s(ws.cell(r, 11).value),
            "keyword_tag": s(ws.cell(r, 13).value),
            "keyword_glossary": s(ws.cell(r, 14).value),
        }
    return out


def load_player_cards_old() -> dict[str, dict]:
    out: dict[str, dict] = {}
    for row in OLD["卡牌"][1:]:
        if len(row) < 6:
            continue
        role = s(row[0])
        name = s(row[1])
        if not name or role not in ("战士", "法老", "恶魔"):
            continue
        out[name] = {
            "role": role,
            "cost": s(row[2]) if len(row) > 2 else "",
            "type": s(row[3]) if len(row) > 3 else "",
            "desc": s(row[5]) if len(row) > 5 else "",
            "rarity": s(row[6]) if len(row) > 6 else "",
            "max_upgrade": s(row[7]) if len(row) > 7 else "",
            "upgrade_per": s(row[8]) if len(row) > 8 else "",
            "upgrade_xp": s(row[9]) if len(row) > 9 else "",
            "old_keyword_tag": s(row[10]) if len(row) > 10 else "",
            "old_keyword_desc": s(row[11]) if len(row) > 11 else "",
        }
    return out


def load_minion_skills_xlsx() -> dict[str, str]:
    ws = openpyxl.load_workbook(XLSX, data_only=True)["小怪设计"]
    out: dict[str, str] = {}
    for r in range(2, ws.max_row + 1):
        name = s(ws.cell(r, 1).value)
        desc = s(ws.cell(r, 8).value) if ws.max_column >= 8 else s(ws.cell(r, 7).value)
        if not name or name.startswith("关于平衡"):
            continue
        if desc:
            out[name] = desc
    return out


def load_minion_skills_old() -> dict[str, str]:
    out: dict[str, str] = {}
    for row in OLD["小怪设计"][2:]:
        if not row:
            continue
        name = s(row[0])
        if not name or name.startswith("关于平衡"):
            continue
        desc = s(row[7]) if len(row) > 7 else ""
        if desc:
            out[name] = desc
    return out


def load_relics_xlsx() -> dict[str, dict]:
    ws = openpyxl.load_workbook(XLSX, data_only=True)["遗物"]
    out: dict[str, dict] = {}
    for r in range(3, ws.max_row + 1):
        rid = s(ws.cell(r, 1).value)
        if not rid or rid == "遗物ID":
            continue
        out[rid] = {
            "name": s(ws.cell(r, 2).value),
            "rarity": s(ws.cell(r, 3).value),
            "category": s(ws.cell(r, 4).value),
            "desc": s(ws.cell(r, 5).value).replace("\n", " "),
            "source": s(ws.cell(r, 6).value).replace("\n", " "),
            "growth": s(ws.cell(r, 7).value) if ws.max_column >= 7 else "",
        }
    return out


def load_relics_old() -> dict[str, dict]:
    out: dict[str, dict] = {}
    for row in OLD["遗物"][2:]:
        if not row or not row[0] or row[0] == "遗物ID":
            continue
        rid = s(row[0])
        out[rid] = {
            "name": s(row[1]),
            "desc": s(row[4]).replace("\n", " ") if len(row) > 4 else "",
            "growth": s(row[5]).replace("\n", " ") if len(row) > 5 else "",
        }
    return out


def load_encounters_xlsx() -> dict[str, str]:
    ws = openpyxl.load_workbook(XLSX, data_only=True)["怪物组合"]
    out: dict[str, str] = {}
    for r in range(2, ws.max_row + 1):
        key = s(ws.cell(r, 1).value) or s(ws.cell(r, 2).value)
        xp = s(ws.cell(r, 6).value)
        if key and xp and key != "敌方阵容（前+中+后）":
            out[key] = xp
    return out


def load_encounters_old() -> dict[str, str]:
    out: dict[str, str] = {}
    for row in OLD["怪物组合"][1:]:
        if not row:
            continue
        key = s(row[0]) or s(row[1] if len(row) > 1 else "")
        xp = s(row[5]) if len(row) > 5 else ""
        if key and xp:
            out[key] = xp
    return out


def load_boss_cards_xlsx() -> dict[str, str]:
    ws = openpyxl.load_workbook(XLSX, data_only=True)["Boss设计"]
    out: dict[str, str] = {}
    for r in range(2, ws.max_row + 1):
        card = s(ws.cell(r, 1).value)
        desc = s(ws.cell(r, 8).value)
        if card and desc and card not in ("Boss基本数值",):
            out[card] = desc
    return out


def load_boss_cards_old() -> dict[str, str]:
    out: dict[str, str] = {}
    for row in OLD["Boss设计"][2:]:
        if not row:
            continue
        card = s(row[0])
        desc = s(row[7]) if len(row) > 7 else ""
        if card and desc:
            out[card] = desc
    return out


def load_talents_xlsx() -> list[dict]:
    ws = openpyxl.load_workbook(XLSX, data_only=True)["角色"]
    out: list[dict] = []
    for r in range(4, ws.max_row + 1):
        lvl = s(ws.cell(r, 1).value)
        if not lvl.isdigit():
            continue
        out.append(
            {
                "level": int(lvl),
                "knight_hp": s(ws.cell(r, 2).value),
                "knight_xp": s(ws.cell(r, 3).value),
                "knight_spd": s(ws.cell(r, 4).value),
                "mage_hp": s(ws.cell(r, 5).value),
                "mage_xp": s(ws.cell(r, 6).value),
                "mage_spd": s(ws.cell(r, 7).value),
                "ranger_hp": s(ws.cell(r, 8).value),
                "ranger_xp": s(ws.cell(r, 9).value),
                "ranger_spd": s(ws.cell(r, 10).value),
                "knight_t1": s(ws.cell(r, 14).value),
                "knight_t2": s(ws.cell(r, 15).value),
                "mage_t1": s(ws.cell(r, 17).value),
                "mage_t2": s(ws.cell(r, 18).value),
                "ranger_t1": s(ws.cell(r, 20).value),
                "ranger_t2": s(ws.cell(r, 21).value),
            }
        )
    return out


def load_talents_old() -> list[dict]:
    out: list[dict] = []
    for row in OLD["角色"][3:]:
        lvl = s(row[0]) if row else ""
        if not lvl.isdigit():
            continue
        out.append(
            {
                "level": int(lvl),
                "knight_hp": s(row[1]),
                "knight_spd": s(row[4]) if len(row) > 4 else "",
                "mage_hp": s(row[5]) if len(row) > 5 else "",
                "mage_spd": s(row[8]) if len(row) > 8 else "",
                "ranger_hp": s(row[9]) if len(row) > 9 else "",
                "ranger_spd": s(row[12]) if len(row) > 12 else "",
                "knight_t1": s(row[16]) if len(row) > 16 else "",
                "knight_t2": s(row[17]) if len(row) > 17 else "",
                "mage_t1": s(row[19]) if len(row) > 19 else "",
                "mage_t2": s(row[20]) if len(row) > 20 else "",
                "ranger_t1": s(row[22]) if len(row) > 22 else "",
                "ranger_t2": s(row[23]) if len(row) > 23 else "",
            }
        )
    return out


def parse_code_relics() -> set[str]:
    text = (ROOT / "Scripts" / "Expedition" / "RelicDatabase.cs").read_text(encoding="utf-8")
    id_map = {
        "SunPyramid": "sun_pyramid",
        "KnightInCastle": "knight_in_castle",
        "BloodAlter": "blood_alter",
        "JadeStone": "jade_stone",
        "JadeRing": "jade_ring",
        "JadeDagger": "jade_dagger",
        "BurningBoots": "burning_boots",
        "CrimsonBurningBoots": "crimson_burning_boots",
        "FlameSword": "flame_sword",
        "IronArmor": "iron_armor",
        "WarriorHelmet": "warrior_helmet",
        "CatStatue": "cat_statue",
        "ElfBow": "elf_bow",
        "DragonRing": "dragon_ring",
        "PaladinShield": "paladin_shield",
        "SilverMoonPendant": "silver_moon_pendant",
        "TaichiRing": "taichi_ring",
        "LeafOfMiracle": "leaf_of_miracle",
        "Bonfire": "bonfire",
    }
    ids: set[str] = set()
    for m in re.findall(r'RelicIds\.(\w+)|"([a-z][a-z0-9_]+)"', text):
        token = m[0] or m[1]
        ids.add(id_map.get(token, re.sub(r"(?<!^)(?=[A-Z])", "_", token).lower()))
    return ids


def main() -> None:
    lines: list[str] = [
        "# Grimhand v0.8 总览表 → 代码实现清单",
        "",
        "对比基准：桌面 `Grimhand实际内容总览表v0.8.xlsx` vs 仓库 `_v08_excel_full.json`（旧快照）+ 当前代码。",
        "",
        "---",
        "",
    ]

    # Global balance rule
    lines += [
        "## 0. 全局平衡规则（优先改底层公式）",
        "",
        "| 项目 | 旧 (v0.7) | 新 (v0.8) | 影响代码 |",
        "|------|-----------|-----------|----------|",
        "| 小怪缩放 | 伤害每2层+1，护甲每3层+1，HP每层+2 | 伤害每3层+1，护甲每5层+1，HP每层+1.5 | `MonsterScalingRules` / 遭遇生成 / 层数缩放逻辑 |",
        "| 基础伤害 | ×1.0 | ×0.6（v0.8平衡版） | 全部怪物/Boss 卡牌 `Card_m_*` 数值、Boss AOE |",
        "| 角色属性表 | 仅 HP+SPD，无局内升级经验列 | 新增「升级用经验」列（战士/法老/恶魔各一列） | 局内升级经验曲线（若已实现需对齐） |",
        "",
    ]

    # Player cards
    new_pc = load_player_cards_xlsx()
    old_pc = load_player_cards_old()
    card_changes: list[str] = []
    for name in sorted(set(new_pc) | set(old_pc)):
        if name not in old_pc:
            card_changes.append(f"- **新增** `{name}`（{new_pc[name]['role']}）：{new_pc[name]['desc']}")
            continue
        if name not in new_pc:
            card_changes.append(f"- **删除** `{name}`")
            continue
        n, o = new_pc[name], old_pc[name]
        fields = []
        for key, label in [
            ("cost", "费用"),
            ("type", "类型"),
            ("desc", "效果描述"),
            ("rarity", "稀有度"),
            ("max_upgrade", "可升级次数"),
            ("upgrade_per", "每次升级"),
            ("upgrade_xp", "升级经验"),
            ("special_effect", "特别效果"),
        ]:
            if s(n.get(key)) != s(o.get(key)):
                fields.append(f"{label} `{o.get(key)}` → `{n.get(key)}`")
        if fields:
            card_changes.append(f"- **`{name}`**（{n['role']}）：" + "；".join(fields))

    lines += [
        f"## 1. 玩家卡牌（共 {len(new_pc)} 张，实质变更 {len(card_changes)} 项）",
        "",
        "涉及文件：`Card_*.asset`、`CardDescriptionCatalog.cs`、`CardUpgradeCatalog.cs`、`KeywordCatalog.cs`、`SpecialCardRules.cs`",
        "",
    ]
    lines.extend(card_changes or ["- （无实质数值/描述变更）"])
    lines.append("")

    # Schema note for keywords
    lines += [
        "### 1.1 表结构变化（关键词列）",
        "",
        "v0.8 将「关键词说明」拆成两列：",
        "- **特别效果**（col11）：该卡专属机制，如【强固×层数】【防御架势】",
        "- **关键词**（col13）+ **描述**（col14）：词条 glossary，如【应对攻击】的通用解释",
        "",
        "代码需同步：`KeywordCatalog`（通用词条）+ 卡牌 `Keywords[]` / `SpecialCardRules`（卡面机制）+ `CardDescriptionCatalog`（UI 文案）",
        "",
    ]

    # Relics
    new_r = load_relics_xlsx()
    old_r = load_relics_old()
    code_r = parse_code_relics()
    lines += ["## 2. 遗物", "", "涉及文件：`RelicDatabase.cs`、`RelicIds.cs`、`RelicEffectRules.cs`、`RelicBattleRules.cs`、立绘 PNG", ""]
    for rid in sorted(set(new_r) - set(old_r)):
        r = new_r[rid]
        lines.append(f"- **新增** `{rid}` **{r['name']}**（{r['category']}）：{r['desc']}")
        lines.append(f"  - 获取：{r['source']}；成长：{r['growth'] or '-'}")
        lines.append(f"  - 代码：**未实现**")
    for rid in sorted(set(new_r) & set(old_r)):
        nr, orow = new_r[rid], old_r[rid]
        if nr["desc"] != orow["desc"] or nr.get("growth") != orow.get("growth"):
            parts = []
            if nr["desc"] != orow["desc"]:
                parts.append(f"效果 `{orow['desc']}` → `{nr['desc']}`")
            if nr.get("growth") != orow.get("growth"):
                parts.append(f"成长 `{orow.get('growth')}` → `{nr.get('growth')}`")
            lines.append(f"- **`{rid}`** {nr['name']}：" + "；".join(parts))
    for rid in sorted(set(new_r) - code_r):
        if rid in new_r:
            lines.append(f"- **代码缺口**：`{rid}` 在 Excel 有定义，但 `RelicDatabase` 未注册")
    lines.append("")

    # Minions
    new_m = load_minion_skills_xlsx()
    old_m = load_minion_skills_old()
    m_changes = [name for name in sorted(set(new_m) & set(old_m)) if new_m[name] != old_m[name]]
    lines += [
        f"## 3. 小怪技能牌（{len(m_changes)} 张描述/数值变更）",
        "",
        "涉及文件：对应 `Card_m_*` / `Card_g_*` 资产、`MonsterContentGenerator`、`MinionTraitCatalog`",
        "",
        "**缩放公式变更**见第 0 节；下列为表内基础描述变更（已含 ×0.6 伤害调整）：",
        "",
    ]
    for name in m_changes:
        lines.append(f"- **`{name}`**：`{old_m[name]}` → `{new_m[name]}`")
    lines.append("")

    # Boss cards
    new_b = load_boss_cards_xlsx()
    old_b = load_boss_cards_old()
    b_changes = {k: (old_b[k], new_b[k]) for k in set(new_b) & set(old_b) if new_b[k] != old_b[k]}
    lines += ["## 4. Boss 技能牌", ""]
    for k, (o, n) in sorted(b_changes.items()):
        lines.append(f"- **`{k}`**：`{o}` → `{n}`")
    lines.append("")

    # Encounters XP
    new_e = load_encounters_xlsx()
    old_e = load_encounters_old()
    e_changes = {k: (old_e[k], new_e[k]) for k in set(new_e) & set(old_e) if new_e[k] != old_e[k]}
    lines += [
        f"## 5. 怪物组合 / 经验奖励（{len(e_changes)} 项）",
        "",
        "涉及文件：`MonsterEncounterCatalog.cs`、远征战斗结算经验逻辑",
        "",
    ]
    for k, (o, n) in sorted(e_changes.items()):
        lines.append(f"- **{k}**：经验 `{o}` → `{n}`")
    lines.append("")

    # Icons
    lines += [
        "## 6. 新增「icon设计」Sheet（状态立绘图标）",
        "",
        "涉及文件：`BattleUiIconCatalogSO`、`CombatantPortraitView` 状态条、美术资源",
        "",
        "| Icon ID | 状态 | 用法 |",
        "|---------|------|------|",
        "| armoracq_down | 破损 | 立绘下方显示层数 |",
        "| armoracq_up | 强固 | 立绘下方显示层数 |",
        "| damage_down | 虚弱 | 立绘下方显示层数 |",
        "| damage_up | 增伤 | 立绘下方显示层数 |",
        "| defense_down | 易伤 | 立绘下方显示层数 |",
        "| defense_up | 减伤 | 立绘下方显示层数 |",
        "| spd_down | 减速 | 立绘下方显示层数 |",
        "| spd_up | 加速 | 立绘下方显示层数 |",
        "",
    ]

    # Talents / roles
    new_t = {t["level"]: t for t in load_talents_xlsx()}
    old_t = {t["level"]: t for t in load_talents_old()}
    lines += ["## 7. 角色属性 / 天赋表", ""]
    lines.append("### 7.1 局内等级属性（HP/SPD/升级经验）")
    lines.append("")
    for lvl in range(1, 11):
        if lvl not in new_t:
            continue
        nt, ot = new_t[lvl], old_t.get(lvl, {})
        for char, hp_k, spd_k, xp_k in [
            ("战士", "knight_hp", "knight_spd", "knight_xp"),
            ("法老", "mage_hp", "mage_spd", "mage_xp"),
            ("恶魔", "ranger_hp", "ranger_spd", "ranger_xp"),
        ]:
            parts = []
            if s(nt.get(hp_k)) != s(ot.get(hp_k)):
                parts.append(f"HP {ot.get(hp_k,'?')}→{nt.get(hp_k)}")
            if s(nt.get(spd_k)) != s(ot.get(spd_k)):
                parts.append(f"SPD {ot.get(spd_k,'?')}→{nt.get(spd_k)}")
            if xp_k in nt and s(nt.get(xp_k)) and s(nt.get(xp_k)) != s(ot.get(xp_k)):
                parts.append(f"升级经验 → {nt.get(xp_k)}")
            if parts:
                lines.append(f"- Lv{lvl} {char}：" + "，".join(parts))
    lines.append("")
    lines.append("### 7.2 局外天赋节点")
    lines.append("")
    lines.append("涉及文件：`TalentCatalog.cs`")
    lines.append("")
    for lvl in range(1, 11):
        if lvl not in new_t or lvl not in old_t:
            continue
        nt, ot = new_t[lvl], old_t[lvl]
        for char, t1k, t2k in [
            ("战士", "knight_t1", "knight_t2"),
            ("法老", "mage_t1", "mage_t2"),
            ("恶魔", "ranger_t1", "ranger_t2"),
        ]:
            for slot, key in [(1, t1k), (2, t2k)]:
                if s(nt.get(key)) and s(nt.get(key)) != s(ot.get(key)):
                    lines.append(f"- Lv{lvl} {char} 槽{slot}：`{ot.get(key)}` → `{nt.get(key)}`")
    lines.append("")

    # Events / consumables unchanged
    lines += [
        "## 8. 事件 / 消耗品",
        "",
        "与仓库 `_v08_excel_full.json` **无差异**（本次 v0.8 表未改这两张 sheet）。若代码尚未完全对齐旧表，以 Excel 当前内容为准单独审计。",
        "",
    ]

    # Implementation order
    lines += [
        "## 9. 建议实施顺序",
        "",
        "1. **全局缩放公式**（第 0 节）→ 一次改完所有层数相关数值",
        "2. **小怪 + Boss 卡牌资产**（第 3–4 节）→ 可脚本批量同步",
        "3. **玩家卡牌**（第 1 节）→ 资产 + 描述 Catalog + 升级 Catalog + 特例规则",
        "4. **遗物**（第 2 节）→ 含新增 `holysun_spellbook` 机制实现",
        "5. **遭遇经验**（第 5 节）",
        "6. **角色/天赋**（第 7 节）",
        "7. **状态图标 UI**（第 6 节）",
        "",
        "---",
        "",
        f"*自动生成；详细 diff 见 `Docs/_v08_diff_report.txt`*",
    ]

    OUT.write_text("\n".join(lines), encoding="utf-8")
    print(f"Wrote {OUT}")
    print(f"Player card changes: {len(card_changes)}")
    print(f"Minion skill changes: {len(m_changes)}")
    print(f"Encounter XP changes: {len(e_changes)}")


if __name__ == "__main__":
    main()
