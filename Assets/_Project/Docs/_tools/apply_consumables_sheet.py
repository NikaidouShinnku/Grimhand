#!/usr/bin/env python3
"""Write consumables sheet in Grimhand实际卡牌遗物总览表.xlsx.

By default reads/writes a copy under Assets/_Project/Docs/ only.
Pass an explicit --xlsx path if you need to update another file (e.g. your local overview).
"""
from __future__ import annotations

import argparse
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font

REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_XLSX = REPO_ROOT / "Docs" / "Grimhand实际卡牌遗物总览表.xlsx"

GENERAL_RULE = (
    "【通用规则】远征栏位5个、不堆叠；满栏需替换。"
    "战斗内仅规划阶段可用，不消耗能量。"
    "每场不限使用次数（可连续使用多个不同消耗品）；"
    "每个消耗品使用后即从栏位消失。"
)

HEADERS = ["消耗品ID", "名称", "效果描述", "获取途径", "备注"]

ROWS = [
    [
        "small_healing_potion",
        "小治疗药水",
        "回复该角色15 HP（含遗物治疗加成）",
        "宝箱 / 战后奖励 / 商店",
        "",
    ],
    [
        "large_healing_potion",
        "大治疗药水",
        "全队存活成员各回复10 HP",
        "宝箱 / 战后奖励 / 商店",
        "",
    ],
    [
        "strength_potion",
        "力量药剂",
        "本回合该角色ATK+30%",
        "宝箱 / 战后奖励 / 商店",
        "",
    ],
    [
        "ironskin_potion",
        "铁壁药剂",
        "本回合该角色DEF+30%",
        "宝箱 / 战后奖励 / 商店",
        "",
    ],
    [
        "spring_bottle",
        "泉水瓶",
        "回复该角色15 HP",
        "魔法泉水事件（装瓶带走）",
        "效果同小治疗药水，主要来源为事件",
    ],
    [
        "mirror_shard",
        "镜之碎片",
        "复制上一回合最后打出的己方攻击牌，由原出牌者再执行一次完整效果",
        "镜中幻影事件（打碎镜子）",
        "第1回合不可用；上回合须打出攻击牌；原出牌者须存活",
    ],
    [
        "scroll_page",
        "古卷残页",
        "本回合能量+2（不超过能量上限）",
        "被诅咒的书架事件（撕页带走）",
        "",
    ],
    [
        "smoke_bomb",
        "烟雾弹",
        "本回合所有角色（含敌我）闪避率+50%",
        "商店 / 宝箱 / 战后奖励",
        "",
    ],
]

WIDTHS = [22, 14, 52, 28, 36]


def apply(xlsx_path: Path) -> None:
    if not xlsx_path.is_file():
        raise FileNotFoundError(xlsx_path)

    wb = openpyxl.load_workbook(xlsx_path)
    if "消耗品" in wb.sheetnames:
        ws = wb["消耗品"]
        if ws.max_row:
            ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet("消耗品")

    rule_cell = ws.cell(1, 1, GENERAL_RULE)
    rule_cell.font = Font(bold=True)
    rule_cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))

    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(2, col, header)
        cell.font = Font(bold=True)

    for row_idx, row in enumerate(ROWS, 3):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row_idx, col_idx, value)

    for idx, width in enumerate(WIDTHS, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(HEADERS)):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(xlsx_path)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX, help="Target workbook path")
    args = parser.parse_args()
    apply(args.xlsx)
