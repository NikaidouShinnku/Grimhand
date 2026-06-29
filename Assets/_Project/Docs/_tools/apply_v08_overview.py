#!/usr/bin/env python3
"""Apply Grimhand v0.8 overview xlsx → Unity assets + generated C# catalogs."""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
XLSX = Path(r"c:\Users\Kelthuzad\Desktop\Grimhand实际内容总览表v0.8.xlsx")
CARDS_DIR = ROOT / "Data" / "Cards"
DESC_CS = ROOT / "Scripts" / "Content" / "CardDescriptionCatalog.cs"
UPGRADE_CS = ROOT / "Scripts" / "Core" / "CardUpgradeCatalog.cs"
KEYWORD_CS = ROOT / "Scripts" / "Battle" / "Rules" / "KeywordCatalog.cs"
JSON_OUT = ROOT / "Docs" / "_v08_excel_authoritative.json"

NAME_ALIASES = {"日光审判": "太阳审判"}

RE_DAMAGE = re.compile(r"造成\s*(\d+)\s*点?伤害")
RE_BLOCK = re.compile(r"(?:获得|添加|为.*?添加|各获得)\s*(\d+)\s*(?:点)?护甲")
RE_HEAL = re.compile(r"(?:治疗|回复)\s*(?:一名队友|自己|等量)?\s*(\d+)\s*HP", re.I)
RE_SACRIFICE_HP = re.compile(r"献祭\s*(\d+)\s*HP")
RE_SACRIFICE_PCT = re.compile(r"献祭\s*(\d+)%")
RE_BONUS_HP = re.compile(r"额外\s*[+＋]\s*(\d+)")
RE_BONUS_DMG = re.compile(r"伤害[+＋]\s*(\d+)")
RE_IGNORE_BLOCK = re.compile(r"无视(?:目标)?\s*(\d+)%?\s*护甲")
RE_IGNORE_ARMOR_FULL = re.compile(r"无视护甲")
RE_HIT_COUNT = re.compile(r"重复\s*(\d+)\s*次")
RE_RESPOND_MIT = re.compile(r"获得\s*(\d+)%\s*减伤")
RE_SPLASH = re.compile(r"(\d+)%\s*的?伤害")
RE_LIFESTEAL = re.compile(r"回复(?:造成)?伤害\s*(\d+)%")
RE_LIFESTEAL_FULL = re.compile(r"回复等量HP")
RE_ON_KILL_HEAL = re.compile(r"击杀(?:回复|恢复)\s*(\d+)\s*HP")

SKIP_KEYWORD_LABELS = {
    "卡牌颜色", "白", "绿", "蓝", "紫", "橙/金", "橙金", "传说", "普通", "稀有", "史诗", "超稀有",
    "描述", "初始牌组", "卡牌名称", "数量", "角色",
}

KW_ID_MAP = {
    "消耗": "exhaust", "献祭": "sacrifice", "应对攻击": "parry",
    "应对防御": "respond_defense", "应对状态": "respond_status",
    "位置": "position", "AOE": "aoe", "中毒": "poison", "灼烧": "burn",
    "减速": "slow", "污染": "polluted", "召唤": "summon",
    "自毁": "self_destruct", "额外手牌": "bonus_hand", "虚化": "ethereal",
    "强固": "armor_up", "破损": "armor_down", "增伤": "damage_up",
    "虚弱": "weaken", "易伤": "vulnerable", "减伤": "damage_reduction",
}


def cell_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def load_workbook_data() -> dict[str, list[list[str]]]:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    data: dict[str, list[list[str]]] = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows: list[list[str]] = []
        for r in range(1, ws.max_row + 1):
            row = [cell_str(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
            while row and row[-1] == "":
                row.pop()
            if any(x for x in row):
                rows.append(row)
        data[name] = rows
    return data


def decode_display(raw: str) -> str:
    if "\\u" in raw:
        try:
            return raw.encode("utf-8").decode("unicode_escape")
        except UnicodeDecodeError:
            pass
    return raw


def index_assets() -> dict[str, dict]:
    by_name: dict[str, dict] = {}
    by_id: dict[str, dict] = {}
    for path in CARDS_DIR.glob("Card_*.asset"):
        text = path.read_text(encoding="utf-8")
        m_id = re.search(r"CardId:\s*(\S+)", text)
        m_name = re.search(r'DisplayName:\s*"([^"]*)"', text)
        if not m_id:
            continue
        cid = m_id.group(1)
        name = decode_display(m_name.group(1)) if m_name else cid
        entry = {"path": path, "id": cid, "name": name}
        by_id[cid] = entry
        by_name[name] = entry
    return {"name": by_name, "id": by_id}


def strip_brackets(s: str) -> str:
    return re.sub(r"[【】\[\]]", "", (s or "").strip())


def keyword_id(label: str) -> str | None:
    label = strip_brackets(label)
    if not label or label in SKIP_KEYWORD_LABELS:
        return None
    if "×" in label or "层数" in label:
        base = re.sub(r"[×x].*", "", label).strip()
        return KW_ID_MAP.get(base)
    if label in KW_ID_MAP:
        return KW_ID_MAP[label]
    if re.match(r"^[前中后/\\、]+$", label):
        return None
    return None


def extract_keywords_from_desc(desc: str) -> list[str]:
    ids: list[str] = []
    seen: set[str] = set()
    for raw in re.findall(r"【([^】]+)】", desc or ""):
        label = raw.strip()
        if re.match(r"^[前中后/、\\]+$", label):
            continue
        kid = keyword_id(label)
        if kid and kid not in seen:
            seen.add(kid)
            ids.append(kid)
    return ids


def parse_upgrade_effect(effect: str) -> dict:
    effect = (effect or "").strip()
    if not effect or effect == "-":
        return {}
    if "额外重复" in effect:
        return {"raw": effect}
    m = re.search(r"[+＋](\d+)\s*伤害", effect)
    if m and "目标受到" not in effect:
        return {"damagePerLevel": int(m.group(1))}
    m = re.search(r"[+＋](\d+)\s*护甲", effect)
    if m:
        return {"blockPerLevel": int(m.group(1))}
    m = re.search(r"[+＋](\d+)\s*恢复?HP", effect, re.I)
    if m:
        return {"healPerLevel": int(m.group(1))}
    m = re.search(r"[+＋](\d+)\s*HP", effect, re.I)
    if m:
        return {"healPerLevel": int(m.group(1))}
    m = re.search(r"[+＋](\d+)\s*层中毒", effect)
    if m:
        return {"damagePerLevel": int(re.search(r"[+＋](\d+)\s*伤害", effect).group(1)) if re.search(r"[+＋](\d+)\s*伤害", effect) else 0,
                "poisonStacksPerLevel": int(m.group(1))}
    m = re.search(r"目标受到的伤害[+＋](\d+)", effect)
    if m:
        return {"damagePerLevel": int(re.search(r"[+＋](\d+)\s*伤害", effect).group(1)) if re.search(r"[+＋](\d+)\s*伤害", effect) else 1}
    m = re.search(r"[+＋](\d+)%", effect)
    if m and "回复" in effect:
        return {"healPerLevel": int(m.group(1))}
    return {"raw": effect}


def infer_patches(desc: str) -> dict[str, int | str | bool]:
    p: dict[str, int | str | bool] = {"ScaleWithAttack": 0, "ScaleWithDefense": 0}
    if not desc:
        return p

    hits = list(RE_DAMAGE.finditer(desc))
    if hits:
        p["Value"] = int(hits[-1].group(1))

    m = RE_BLOCK.search(desc)
    if m:
        p["Value"] = int(m.group(1))

    m = RE_HEAL.search(desc)
    if m:
        p["Value"] = int(m.group(1))

    m = RE_SACRIFICE_HP.search(desc)
    if m:
        p["SelfDamageFlat"] = int(m.group(1))

    m = RE_SACRIFICE_PCT.search(desc)
    if m:
        p["HealMaxHpPercent"] = int(m.group(1))

    m = RE_BONUS_HP.search(desc)
    if m:
        p["BonusIfTargetHpBelowFlat"] = int(m.group(1))
        if "HP<" in desc or "HP＜" in desc:
            p["BonusIfTargetHpBelowPercent"] = 50

    m = RE_IGNORE_BLOCK.search(desc)
    if m:
        p["IgnoreDefPercent"] = int(m.group(1))
    elif RE_IGNORE_ARMOR_FULL.search(desc):
        p["IgnoreDefPercent"] = 100

    m = RE_HIT_COUNT.search(desc)
    if m:
        p["HitCount"] = int(m.group(1))

    m = RE_RESPOND_MIT.search(desc)
    if m:
        p["Value"] = int(m.group(1))

    m = RE_SPLASH.search(desc)
    if m and ("身后" in desc or "后方" in desc or "贯通" in desc or "80%" in desc):
        p["SplashBehindTarget"] = 1
        p["SplashPowerPercent"] = int(m.group(1))

    m = RE_LIFESTEAL.search(desc)
    if m:
        p["LifestealPercent"] = int(m.group(1))
    elif RE_LIFESTEAL_FULL.search(desc):
        p["LifestealPercent"] = 100

    m = RE_ON_KILL_HEAL.search(desc)
    if m:
        p["OnKillHealAmount"] = int(m.group(1))

    if "中毒" in desc:
        m = re.search(r"(\d+)\s*层中毒", desc)
        if m:
            p["StatusId"] = "poison"
            p["Stacks"] = int(m.group(1))

    if "减速" in desc:
        m = re.search(r"减速\s*[×x]\s*(\d+)", desc) or re.search(r"(\d+)\s*层减速", desc)
        if m:
            p["StatusId"] = "slow"
            p["Stacks"] = int(m.group(1))

    return p


def patch_field(text: str, field: str, value, *, start: int = 0) -> str:
    if isinstance(value, bool):
        value = 1 if value else 0
    if isinstance(value, str):
        pattern = rf"({re.escape(field)}:\s*)([^\n]*)"
        m = re.search(pattern, text[start:])
        if not m:
            return text
        at = start + m.start()
        return text[:at] + re.sub(pattern, rf"\g<1>{value}", text[at:], count=1)
    pattern = rf"({re.escape(field)}:\s*)(-?\d+)"
    m = re.search(pattern, text[start:])
    if not m:
        return text
    at = start + m.start()
    return text[:at] + re.sub(pattern, rf"\g<1>{value}", text[at:], count=1)


def patch_apply_status_action(text: str, status_id: str, stacks: int) -> str:
    """仅更新 ApplyStatus (Type: 3) 动作块，避免污染伤害/护甲动作。"""
    blocks = list(re.finditer(r"  - Type: 3\n(.*?)(?=\n  - Type:|\n  CardArt:)", text, re.S))
    if not blocks:
        return text
    block = blocks[0]
    segment = block.group(0)
    segment = patch_field(segment, "StatusId", status_id, start=0)
    segment = patch_field(segment, "Stacks", stacks, start=0)
    return text[: block.start()] + segment + text[block.end() :]


def patch_keywords(text: str, desc: str) -> str:
    keywords = extract_keywords_from_desc(desc)
    kw_block = "  Keywords: []" if not keywords else "  Keywords:\n" + "\n".join(f"  - {k}" for k in keywords)
    text = re.sub(r"  Keywords:\s*\n(?:  - .+\n|- .+\n)*", "", text)
    text = re.sub(r"  Keywords: \[\]\n", "", text)
    text = re.sub(r"^- .+\n(?=  Rarity:)", "", text, flags=re.MULTILINE)
    if re.search(r"  Rarity:", text):
        return re.sub(r"(  Rarity: \d+\n)", kw_block + "\n\\1", text, count=1)
    return re.sub(r"(  CardType: \d+\n)", "\\1" + kw_block + "\n", text, count=1)


def apply_patches_to_asset(path: Path, desc: str) -> bool:
    if "剩余所有能量" in desc or "花费剩余所有能量" in desc:
        return False
    text = path.read_text(encoding="utf-8")
    original = text
    text = re.sub(r"ScaleWithAttack:\s*1", "ScaleWithAttack: 0", text)
    text = re.sub(r"ScaleWithDefense:\s*1", "ScaleWithDefense: 0", text)
    patches = infer_patches(desc)
    status_id = patches.pop("StatusId", None)
    stacks = patches.pop("Stacks", None)
    for field, value in patches.items():
        text = patch_field(text, field, value)
    if status_id and stacks is not None:
        text = patch_apply_status_action(text, status_id, stacks)
    new_text = patch_keywords(text, desc)
    if new_text != original:
        path.write_text(new_text, encoding="utf-8")
        return True
    return False


def parse_player_cards(data: dict) -> dict[str, dict]:
    rows = {}
    for row in data["卡牌"][1:]:
        if not row or len(row) < 6:
            continue
        role = row[0]
        name = row[1].strip()
        desc = row[5].strip()
        if role not in ("战士", "法老", "恶魔") or not name or not desc:
            continue
        max_up = row[7] if len(row) > 7 else ""
        eff = row[8].strip() if len(row) > 8 else ""
        xp_raw = row[9].strip() if len(row) > 9 else ""
        xp = 0
        m = re.match(r"(\d+)\s*/\s*级", xp_raw)
        if m:
            xp = int(m.group(1))
        try:
            max_up_i = int(float(max_up)) if max_up not in ("", "-") else 0
        except ValueError:
            max_up_i = 0
        rows[name] = {
            "desc": desc,
            "maxUpgrades": max_up_i,
            "upgradeEffect": eff,
            "xpPerLevel": xp,
            "keywordLabel": row[12].strip() if len(row) > 12 else "",
            "keywordDesc": row[13].strip() if len(row) > 13 else "",
        }
    return rows


def parse_monster_cards(data: dict) -> dict[str, str]:
    rows: dict[str, str] = {}
    in_cards = False
    for row in data["小怪设计"]:
        if not row:
            continue
        if row[0] == "卡牌名称":
            in_cards = True
            continue
        if row[0] == "角色名":
            in_cards = False
            continue
        if not in_cards:
            continue
        name = row[0].strip()
        desc = (row[7] if len(row) > 7 else "").strip()
        if name and desc:
            rows[name] = desc
    return rows


def parse_boss_cards(data: dict) -> dict[str, str]:
    rows: dict[str, str] = {}
    in_cards = False
    for row in data["Boss设计"]:
        if not row:
            continue
        if row[0] == "Boss卡牌":
            in_cards = True
            continue
        if row[0] == "Boss基本数值":
            in_cards = False
            continue
        if not in_cards:
            continue
        name = (row[1] if len(row) > 1 else "").strip()
        desc = (row[7] if len(row) > 7 else "").strip()
        if name and desc and name not in ("卡牌名称",):
            rows[name] = desc
    return rows


def cs_escape(s: str) -> str:
    return s.replace("\\", "\\\\").replace('"', '\\"').replace("\r", " ").replace("\n", " ")


def emit_description_cs(descriptions: dict[str, str], assets: dict) -> None:
    by_name_lines = []
    by_id_lines = []
    for name in sorted(descriptions.keys(), key=lambda x: (x[0], x)):
        desc = cs_escape(descriptions[name])
        by_name_lines.append(f'            ["{cs_escape(name)}"] = "{desc}",')
    for cid, entry in sorted(assets["id"].items()):
        name = entry["name"]
        lookup = NAME_ALIASES.get(name, name)
        if lookup in descriptions:
            by_id_lines.append(
                f'            ["{cid}"] = "{cs_escape(descriptions[lookup])}",'
            )
        elif name in descriptions:
            by_id_lines.append(f'            ["{cid}"] = "{cs_escape(descriptions[name])}",')

    content = f"""using System.Collections.Generic;

namespace Grimhand.Content
{{
    /// <summary>卡牌描述（对照 Grimhand实际内容总览表 v0.8，UI 唯一文案来源）。</summary>
    public static class CardDescriptionCatalog
    {{
        static readonly Dictionary<string, string> ByDisplayName = BuildByName();
        static readonly Dictionary<string, string> ByCardId = BuildById();

        public static bool TryGetByDisplayName(string displayName, out string description)
        {{
            description = null;
            if (string.IsNullOrEmpty(displayName))
                return false;
            return ByDisplayName.TryGetValue(displayName, out description);
        }}

        public static bool TryGetByCardId(string cardId, out string description)
        {{
            description = null;
            if (string.IsNullOrEmpty(cardId))
                return false;
            return ByCardId.TryGetValue(cardId, out description);
        }}

        static Dictionary<string, string> BuildByName() => new()
        {{
{chr(10).join(by_name_lines)}
        }};

        static Dictionary<string, string> BuildById() => new()
        {{
{chr(10).join(by_id_lines)}
        }};
    }}
}}
"""
    DESC_CS.write_text(content, encoding="utf-8")


def emit_upgrade_cs(cards: list[dict]) -> None:
    lines = [
        "using System.Collections.Generic;",
        "",
        "namespace Grimhand.Core",
        "{",
        "    /// <summary>卡牌升级配置（对照 Grimhand实际内容总览表 v0.8 · 卡牌 sheet）。</summary>",
        "    public static class CardUpgradeCatalog",
        "    {",
        "        public sealed class UpgradeSpec",
        "        {",
        "            public int MaxUpgrades { get; set; }",
        "            public int DamagePerLevel { get; set; }",
        "            public int BlockPerLevel { get; set; }",
        "            public int HealPerLevel { get; set; }",
        "            public int CostReductionPerLevel { get; set; }",
        "            public int PoisonStacksPerLevel { get; set; }",
        "            public int SlowStacksPerLevel { get; set; }",
        "            public int XpCostPerLevel { get; set; }",
        "        }",
        "",
        "        static readonly Dictionary<string, UpgradeSpec> ByDisplayName = Build();",
        "",
        "        public static bool TryGetByDisplayName(string displayName, out UpgradeSpec spec)",
        "        {",
        "            if (string.IsNullOrEmpty(displayName))",
        "            {",
        "                spec = null;",
        "                return false;",
        "            }",
        "",
        "            return ByDisplayName.TryGetValue(displayName, out spec);",
        "        }",
        "",
        "        public static bool CanUpgrade(string displayName, int currentLevel) =>",
        "            TryGetByDisplayName(displayName, out var spec) && currentLevel < spec.MaxUpgrades;",
        "",
        "        public static int GetXpCostPerLevel(string displayName) =>",
        "            TryGetByDisplayName(displayName, out var spec) ? spec.XpCostPerLevel : 0;",
        "",
        "        static Dictionary<string, UpgradeSpec> Build() => new()",
        "        {",
    ]
    for c in cards:
        u = parse_upgrade_effect(c["upgradeEffect"])
        lines.append(
            f'            ["{cs_escape(c["name"])}"] = new() {{ MaxUpgrades = {c["maxUpgrades"]}, '
            f'DamagePerLevel = {u.get("damagePerLevel", 0)}, '
            f'BlockPerLevel = {u.get("blockPerLevel", 0)}, '
            f'HealPerLevel = {u.get("healPerLevel", 0)}, '
            f'CostReductionPerLevel = {u.get("costReductionPerLevel", 0)}, '
            f'PoisonStacksPerLevel = {u.get("poisonStacksPerLevel", 0)}, '
            f'SlowStacksPerLevel = {u.get("slowStacksPerLevel", 0)}, '
            f'XpCostPerLevel = {c.get("xpPerLevel", 0)} }},'
        )
    lines.extend(["        };", "    }", "}", ""])
    UPGRADE_CS.write_text("\n".join(lines), encoding="utf-8")


def emit_keyword_cs(keywords: dict[str, str]) -> None:
    lines = [
        "using System.Collections.Generic;",
        "using System.Text;",
        "using Grimhand.Battle.Model;",
        "",
        "namespace Grimhand.Battle.Rules",
        "{",
        "    /// <summary>关键词（对照 Grimhand实际内容总览表 v0.8 · 卡牌 sheet）。</summary>",
        "    public static class KeywordCatalog",
        "    {",
        "        static readonly Dictionary<string, KeywordDefinition> Definitions = BuildDefinitions();",
        "",
        "        public static bool TryGet(string keywordId, out KeywordDefinition definition) =>",
        "            Definitions.TryGetValue(keywordId, out definition);",
        "",
        "        public static string BuildTooltipText(IReadOnlyList<string> keywordIds)",
        "        {",
        "            if (keywordIds == null || keywordIds.Count == 0)",
        "                return \"\";",
        "            var sb = new StringBuilder();",
        "            for (var i = 0; i < keywordIds.Count; i++)",
        "            {",
        "                var id = keywordIds[i];",
        "                if (string.IsNullOrEmpty(id) || !TryGet(id, out var def))",
        "                    continue;",
        "                if (sb.Length > 0) sb.Append(\"\\n\\n\");",
        "                sb.Append(def.DisplayName).Append(\"：\").Append(def.Description);",
        "            }",
        "            return sb.ToString();",
        "        }",
        "",
        "        public static string BuildRichTooltipText(IReadOnlyList<string> keywordIds)",
        "        {",
        "            if (keywordIds == null || keywordIds.Count == 0)",
        "                return \"\";",
        "            var sb = new StringBuilder();",
        "            for (var i = 0; i < keywordIds.Count; i++)",
        "            {",
        "                var id = keywordIds[i];",
        "                if (string.IsNullOrEmpty(id) || !TryGet(id, out var def))",
        "                    continue;",
        "                if (sb.Length > 0) sb.Append(\"\\n\\n\");",
        "                sb.Append(\"<b>\").Append(def.DisplayName).Append(\"</b>\\n\").Append(def.Description);",
        "            }",
        "            return sb.ToString();",
        "        }",
        "",
        "        static Dictionary<string, KeywordDefinition> BuildDefinitions() => new()",
        "        {",
    ]
    seen: set[str] = set()
    for label, desc in sorted(keywords.items(), key=lambda x: x[0]):
        kid = keyword_id(label)
        if not kid or kid in seen:
            continue
        seen.add(kid)
        display = strip_brackets(label) or label
        lines.append(
            f'            ["{kid}"] = new("{kid}", "{cs_escape(display)}", "{cs_escape(desc)}"),'
        )
    lines.extend(["        };", "    }", "}", ""])
    KEYWORD_CS.write_text("\n".join(lines), encoding="utf-8")


def sync_cards(name_to_desc: dict[str, str], assets: dict, label: str) -> tuple[int, list[str]]:
    updated = 0
    missing = []
    for name, desc in name_to_desc.items():
        lookup = NAME_ALIASES.get(name, name)
        entry = assets["name"].get(lookup)
        if not entry:
            missing.append(name)
            continue
        if apply_patches_to_asset(entry["path"], desc):
            updated += 1
            print(f"  [{label}] {name}")
    return updated, missing


def main() -> int:
    sys.stdout.reconfigure(encoding="utf-8")
    data = load_workbook_data()
    JSON_OUT.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    assets = index_assets()
    player = parse_player_cards(data)
    monster = parse_monster_cards(data)
    boss = parse_boss_cards(data)

    all_descriptions: dict[str, str] = {}
    for src in (player, monster, boss):
        for k, v in src.items():
            all_descriptions[k] = v if isinstance(v, str) else v["desc"]

    pu, pm = sync_cards({k: v["desc"] for k, v in player.items()}, assets, "玩家")
    mu, mm = sync_cards(monster, assets, "小怪")
    bu, bm = sync_cards(boss, assets, "Boss")

    upgrade_cards = [
        {
            "name": k,
            "maxUpgrades": v["maxUpgrades"],
            "upgradeEffect": v["upgradeEffect"],
            "xpPerLevel": v["xpPerLevel"],
        }
        for k, v in player.items()
        if v["maxUpgrades"] > 0 and v["upgradeEffect"] not in ("", "-")
    ]
    emit_upgrade_cs(upgrade_cards)
    keywords: dict[str, str] = {}
    for v in player.values():
        lbl = strip_brackets(v["keywordLabel"])
        if lbl and v["keywordDesc"]:
            keywords[lbl] = v["keywordDesc"]
    emit_keyword_cs(keywords)
    emit_description_cs(all_descriptions, assets)

    print(f"\nJSON snapshot: {JSON_OUT}")
    print(f"玩家 {pu}/{len(player)} missing={len(pm)}")
    print(f"小怪 {mu}/{len(monster)} missing={len(mm)}")
    print(f"Boss {bu}/{len(boss)} missing={len(bm)}")
    print(f"Descriptions {len(all_descriptions)} | Upgrades {len(upgrade_cards)} | Keywords {len(keywords)}")
    if pm[:5]:
        print("玩家未匹配:", pm[:10])
    if mm[:5]:
        print("小怪未匹配:", mm[:10])
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
