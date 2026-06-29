#!/usr/bin/env python3
import re
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parents[2]
text = (ROOT / "Scripts/Expedition/TalentCatalog.cs").read_text(encoding="utf-8")
ws = openpyxl.load_workbook(
    Path(r"c:\Users\Kelthuzad\Desktop\Grimhand实际内容总览表v0.8.xlsx"), data_only=True
)["角色"]

char_map = {
    14: ("knight", 1),
    15: ("knight", 2),
    17: ("mage", 1),
    18: ("mage", 2),
    20: ("ranger", 1),
    21: ("ranger", 2),
}
code = {}
pat = re.compile(
    r'Def\((\w+Id),\s*(\d+),\s*(\d+),\s*"[^"]+",\s*"[^"]+",\s*\n\s*"([^"]+)"\)',
    re.MULTILINE,
)
for m in pat.finditer(text):
    cid = m.group(1).replace("Id", "").lower()
    code[(cid, int(m.group(2)), int(m.group(3)))] = m.group(4)

print("=== TALENT excel vs code mismatches ===")
for r in range(4, ws.max_row + 1):
    lvl = ws.cell(r, 1).value
    try:
        ol = int(float(lvl))
    except (TypeError, ValueError):
        continue
    for col, (cid, slot) in char_map.items():
        desc = ws.cell(r, col).value
        if not desc:
            continue
        cd = code.get((cid, slot, ol))
        if cd and str(desc).strip() != cd.strip():
            print(f"Lv{ol} {cid} slot{slot}:")
            print(f"  EXCEL: {desc}")
            print(f"  CODE:  {cd}")
