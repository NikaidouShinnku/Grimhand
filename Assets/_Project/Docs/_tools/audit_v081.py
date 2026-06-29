#!/usr/bin/env python3
"""Diff v0.81 overview (卡牌/遗物/角色) against current in-game implementation."""
from __future__ import annotations

import re
from pathlib import Path

import openpyxl

ROOT = Path(r"c:\Users\Kelthuzad\Documents\GitHub\Grimhand\Grimhand")
SRC = Path(r"c:\Users\Kelthuzad\Desktop\Grimhand实际内容总览表v0.81.xlsx")
OUT = Path(__file__).resolve().parent / "_v081_audit.txt"

CARDS_DIR = ROOT / "Assets" / "_Project" / "Data" / "Cards"
DESC_CATALOG = ROOT / "Assets" / "_Project" / "Scripts" / "Content" / "CardDescriptionCatalog.cs"
RELIC_DB = ROOT / "Assets" / "_Project" / "Scripts" / "Expedition" / "RelicDatabase.cs"
TALENT_CATALOG_DIR = ROOT / "Assets" / "_Project" / "Scripts" / "Battle" / "Rules"  # TalentCatalog lives where?

wb = openpyxl.load_workbook(SRC, data_only=True)

# ---- v0.81 cards ----
ws = wb["卡牌"]
v081_cards = []  # (owner, name, cost, type, rarity, desc)
for row in ws.iter_rows(values_only=True):
    if not row or row[0] is None:
        continue
    owner = str(row[0]).strip()
    name = row[1]
    if not name or not isinstance(name, str):
        continue
    name = name.strip()
    if owner not in ("战士", "法老", "恶魔"):
        continue
    cost = row[2]
    ctype = row[3]
    rarity = row[6]
    desc = row[5]
    v081_cards.append({
        "owner": owner,
        "name": name,
        "cost": "" if cost is None else str(cost),
        "type": "" if ctype is None else str(ctype).strip(),
        "rarity": "" if rarity is None else str(rarity).strip(),
        "desc": "" if desc is None else str(desc).strip(),
    })

# ---- current in-game description catalog (ByDisplayName + ByCardId) ----
text = DESC_CATALOG.read_text(encoding="utf-8")

def extract_entries(body: str) -> dict:
    return dict(re.findall(r'\["([^"]+)"\]\s*=\s*"((?:[^"\\]|\\.)*)"', body))

# find method bodies: BuildByName() => new() { ... };  and  BuildById() => new() { ... };
def find_body(label: str) -> str:
    m = re.search(rf"{label}\(\)\s*=>\s*new\(\)\s*\{{(.*?)\n\s*\}};", text, re.S)
    return m.group(1) if m else ""

cur_by_name = extract_entries(find_body("BuildByName"))
cur_by_id = extract_entries(find_body("BuildById"))

# ---- card assets (YAML) -> cardId, displayName, cost, type, rarity, keywords ----
def parse_card_asset(path: Path):
    t = path.read_text(encoding="utf-8")
    def first(field):
        m = re.search(rf"^\s*{field}:\s*(.+?)$", t, re.M)
        if not m:
            return ""
        v = m.group(1).strip()
        # strip Unity escaped quotes
        v = v.replace('"', '')
        return v
    # Keywords block
    kw = []
    m = re.search(r"Keywords:\s*\n((?:\s+- .+\n|\s*\[\]\n)*)", t)
    if m:
        block = m.group(1)
        if "[]" not in block:
            kw = [x.strip().lstrip("- ").strip() for x in block.splitlines() if x.strip().startswith("-")]
    return {
        "cardId": first("CardId"),
        "displayName": first("DisplayName").encode().decode("unicode_escape") if "\\u" in first("DisplayName") else first("DisplayName"),
        "cost": first("Cost"),
        "cardType": first("CardType"),
        "rarity": first("Rarity"),
        "keywords": kw,
    }

card_assets = {}
for p in CARDS_DIR.glob("Card_*.asset"):
    info = parse_card_asset(p)
    if info["cardId"]:
        card_assets[info["cardId"]] = info

# Map displayName -> cardId
name_to_id = {info["displayName"]: cid for cid, info in card_assets.items()}

# ---- diff cards ----
lines = []
lines.append("### 卡牌 diff (v0.81 vs current)")
lines.append("格式: [v0.81] 名称 | cost/type/rarity | 描述")
lines.append("      [cur ] cardId | desc (ByDisplayName or ByCardId)")
lines.append("")

rarity_map = {"白": "0 Common", "绿": "1 Rare", "蓝": "2 SuperRare", "紫": "3 Epic", "橙": "4 Legendary"}
type_map = {"攻击": "0 Attack", "防御": "1 Defense", "状态": "2 Status"}

for c in v081_cards:
    name = c["name"]
    cur_desc = cur_by_name.get(name)
    card_id = name_to_id.get(name, "")
    asset = card_assets.get(card_id)
    cur_cost = asset["cost"] if asset else "?"
    cur_type = asset["cardType"] if asset else "?"
    cur_rarity = asset["rarity"] if asset else "?"
    cur_kw = asset["keywords"] if asset else []
    desc_match = cur_desc == c["desc"] if cur_desc else False
    cost_match = str(cur_cost) == str(c["cost"]) if cur_cost != "?" else False
    flag = ""
    if not desc_match:
        flag += " DESC"
    if not cost_match:
        flag += " COST"
    if flag or not asset:
        lines.append(f"--- {c['owner']} {name} (id={card_id}){flag} ---")
        lines.append(f"  v0.81: cost={c['cost']} type={c['type']} rarity={c['rarity']}")
        lines.append(f"  v0.81 desc: {c['desc']}")
        lines.append(f"  cur  : cost={cur_cost} type={cur_type} rarity={cur_rarity} kw={cur_kw}")
        lines.append(f"  cur desc : {cur_desc}")
        lines.append("")

# ---- v0.81 relics ----
ws = wb["遗物"]
v081_relics = []
cur_relic_id = None
buf = {}
for row in ws.iter_rows(values_only=True):
    if not row or all(v is None for v in row):
        continue
    rid = row[0]
    if rid and isinstance(rid, str) and re.match(r"^[a-z_]+$", rid.strip()):
        if cur_relic_id:
            v081_relics.append(buf)
        cur_relic_id = rid.strip()
        buf = {"id": cur_relic_id, "name": row[1], "rarity": row[2], "category": row[3],
               "desc_parts": [row[4]] if row[4] else [], "acq": row[5], "growth": row[6]}
    else:
        if not buf:
            continue
        if row[4]:
            buf["desc_parts"].append(row[4])
        if row[5] and not buf.get("acq"):
            buf["acq"] = row[5]
        if row[6] and not buf.get("growth"):
            buf["growth"] = row[6]
if cur_relic_id:
    v081_relics.append(buf)
for r in v081_relics:
    r["desc"] = " ".join(str(x).strip() for x in r["desc_parts"] if x).strip()

# parse RelicDatabase.cs for current relics — match Def(RelicIds.XXX, "name", rarity, "category", "desc", "flag" ...)
relic_text = RELIC_DB.read_text(encoding="utf-8")
cur_relics = {}
# Def(RelicIds.SunPyramid, "太阳金字塔", RelicRarity.Rare, "法老专属", "desc...", "flag"...);
pat = re.compile(
    r'Def\(\s*RelicIds\.(\w+)\s*,\s*"([^"]*)"\s*,\s*RelicRarity\.(\w+)\s*,\s*"([^"]*)"\s*,\s*\n?\s*"((?:[^"\\]|\\.)*)"\s*,\s*\n?\s*"([^"]*)"',
    re.S,
)
for m in pat.finditer(relic_text):
    rid_const = m.group(1)  # e.g. SunPyramid
    # convert PascalCase to snake_case
    rid = re.sub(r"(?<!^)(?=[A-Z])", "_", rid_const).lower()
    cur_relics[rid] = {
        "name": m.group(2),
        "rarity": m.group(3),
        "category": m.group(4),
        "desc": m.group(5),
        "flag": m.group(6),
    }

lines.append("### 遗物 diff (v0.81 vs current RelicDatabase)")
lines.append("")
for r in v081_relics:
    rid = r["id"]
    cur = cur_relics.get(rid)
    cur_desc = cur["desc"] if cur else "<MISSING>"
    cur_name = cur["name"] if cur else "<MISSING>"
    cur_rarity = cur["rarity"] if cur else "?"
    cur_cat = cur["category"] if cur else "?"
    cur_flag = cur["flag"] if cur else "?"
    if cur_desc != r["desc"] or cur_name != r["name"] or cur_cat != r["category"]:
        lines.append(f"--- {rid} ({r['name']}) ---")
        lines.append(f"  v0.81: rarity={r['rarity']} cat={r['category']}")
        lines.append(f"  v0.81 desc: {r['desc']}")
        lines.append(f"  cur  : rarity={cur_rarity} cat={cur_cat} flag={cur_flag}")
        lines.append(f"  cur desc : {cur_desc}")
        lines.append("")

# ---- v0.81 talents (角色 sheet) ----
# Talent columns: M=12 局外等级, N=13 战1, O=14 战2, P=15 局外等级, Q=16 法1, R=17 法2, S=18 局外等级, T=19 恶1, U=20 恶2
ws = wb["角色"]
talent_rows = []
talent_col_indices = {"战1": 13, "战2": 14, "法1": 16, "法2": 17, "恶1": 19, "恶2": 20}
for row in ws.iter_rows(values_only=True):
    if not row:
        continue
    lvl = row[0]
    # only rows where col A is an integer level (talent rows)
    if lvl is None or not isinstance(lvl, int):
        continue
    talents = []
    for lab, idx in talent_col_indices.items():
        if idx >= len(row):
            continue
        v = row[idx]
        if v and isinstance(v, str) and v.strip():
            talents.append(f"{lab}: {v.strip()}")
    if talents:
        talent_rows.append((lvl, talents))

lines.append("### 天赋 v0.81 (按局外等级)")
lines.append("")
for lvl, ts in talent_rows:
    lines.append(f"Lv{lvl}: " + " | ".join(ts))

OUT.write_text("\n".join(lines), encoding="utf-8")
print(f"Audit written to {OUT}")
print(f"  v0.81 cards: {len(v081_cards)}, relics: {len(v081_relics)}, talent rows: {len(talent_rows)}")
