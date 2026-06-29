#!/usr/bin/env python3
"""Compare Desktop v0.8 overview xlsx vs repo snapshot and codebase."""
from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl

XLSX = Path(r"c:\Users\Kelthuzad\Desktop\Grimhand实际内容总览表v0.8.xlsx")
ROOT = Path(__file__).resolve().parents[2]
OLD = ROOT / "Docs" / "_v08_excel_full.json"
OUT_JSON = ROOT / "Docs" / "_v08_diff_report.json"
OUT_TXT = ROOT / "Docs" / "_v08_diff_report.txt"


def cell_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def load_xlsx(path: Path) -> dict[str, list[list[str]]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    data: dict[str, list[list[str]]] = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows: list[list[str]] = []
        for r in range(1, ws.max_row + 1):
            row = [cell_str(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
            while row and row[-1] == "":
                row.pop()
            if any(x for x in row):
                rows.append(row)
        data[name] = rows
    return data


def norm(s) -> str:
    s = cell_str(s)
    s = s.replace("\r\n", "\n").replace("\r", "")
    s = re.sub(r"\s+", " ", s)
    return s.replace("％", "%").replace("＋", "+").replace("—", "-")


def key_for_row(sheet: str, row: list[str]) -> str | None:
    if not row:
        return None
    if sheet == "卡牌":
        return norm(row[1]) if len(row) > 1 and row[1] else norm(row[0]) or None
    if sheet == "遗物":
        k = norm(row[0])
        if k and k not in ("遗物ID",):
            return k
        return norm(row[1]) if len(row) > 1 else None
    if sheet in ("Boss设计", "小怪设计", "怪物组合", "消耗品", "事件"):
        return norm(row[0]) or (norm(row[1]) if len(row) > 1 else None)
    if sheet == "角色":
        k = norm(row[0])
        return f"lv{k}" if k.isdigit() else None
    return norm(row[0]) or None


def diff_sheet(sheet: str, new_rows: list, old_rows: list) -> dict:
    diff = {"added_rows": [], "removed_rows": [], "changed_rows": []}
    new_map: dict[str, list] = {}
    old_map: dict[str, list] = {}
    for row in new_rows:
        k = key_for_row(sheet, row)
        if k:
            new_map[k] = row
    for row in old_rows:
        k = key_for_row(sheet, row)
        if k:
            old_map[k] = row

    for k in sorted(set(new_map) - set(old_map)):
        diff["added_rows"].append({"key": k, "row": new_map[k]})
    for k in sorted(set(old_map) - set(new_map)):
        diff["removed_rows"].append({"key": k, "row": old_map[k]})
    for k in sorted(set(new_map) & set(old_map)):
        nr, orow = new_map[k], old_map[k]
        m = max(len(nr), len(orow))
        nr2 = nr + [""] * (m - len(nr))
        or2 = orow + [""] * (m - len(orow))
        changes = [
            {"col": ci, "old": b, "new": a}
            for ci, (a, b) in enumerate(zip(nr2, or2))
            if norm(a) != norm(b)
        ]
        if changes:
            diff["changed_rows"].append({"key": k, "changes": changes, "new_row": nr, "old_row": orow})
    return diff


def parse_relic_ids(text: str) -> set[str]:
    id_map = {
        "SunPyramid": "sun_pyramid",
        "KnightInCastle": "knight_in_castle",
        "BloodAlter": "blood_alter",
        "JadeStone": "jade_stone",
        "JadeRing": "jade_ring",
        "JadeDagger": "jade_dagger",
        "BurningBoots": "burning_boots",
        "CrimsonBurningBoots": "crimson_burning_boots",
        "FlameSword": "flame_sword",
        "IronArmor": "iron_armor",
        "WarriorHelmet": "warrior_helmet",
        "CatStatue": "cat_statue",
        "ElfBow": "elf_bow",
        "DragonRing": "dragon_ring",
        "PaladinShield": "paladin_shield",
        "SilverMoonPendant": "silver_moon_pendant",
        "TaichiRing": "taichi_ring",
        "LeafOfMiracle": "leaf_of_miracle",
    }
    out: set[str] = set()
    for m in re.findall(r"RelicIds\.(\w+)", text):
        out.add(id_map.get(m, re.sub(r"(?<!^)(?=[A-Z])", "_", m).lower()))
    return out


def index_card_assets() -> dict[str, str]:
    by_name: dict[str, str] = {}
    for p in (ROOT / "Data" / "Cards").glob("Card_*.asset"):
        text = p.read_text(encoding="utf-8")
        m_id = re.search(r"CardId:\s*(\S+)", text)
        m_name = re.search(r'DisplayName:\s*"([^"]*)"', text)
        if not m_name:
            continue
        name = m_name.group(1)
        if "\\u" in name:
            try:
                name = name.encode("utf-8").decode("unicode_escape")
            except UnicodeDecodeError:
                pass
        by_name[name] = m_id.group(1) if m_id else p.stem
    return by_name


def col_label(sheet: str, col: int) -> str:
    labels = {
        "卡牌": {0: "ID", 1: "名称", 2: "角色", 3: "费用", 4: "类型", 5: "描述", 6: "升级描述", 12: "关键词"},
        "遗物": {0: "遗物ID", 1: "名称", 2: "稀有度", 3: "归属", 4: "效果描述", 5: "成长"},
        "事件": {0: "事件ID", 1: "名称", 2: "类型", 3: "选项A", 4: "选项B", 5: "说明"},
        "消耗品": {0: "ID", 1: "名称", 2: "效果"},
        "小怪设计": {0: "怪物ID", 1: "名称", 2: "HP", 3: "SPD", 4: "特性", 5: "技能池"},
        "Boss设计": {0: "BossID", 1: "名称", 2: "HP", 3: "SPD", 4: "机制", 5: "牌组"},
        "怪物组合": {0: "遭遇ID", 1: "楼层", 2: "组合"},
    }
    return labels.get(sheet, {}).get(col, f"col{col}")


def format_change(sheet: str, item: dict) -> list[str]:
    lines = [f"  [{item['key']}]"]
    for ch in item["changes"]:
        label = col_label(sheet, ch["col"])
        old = (ch["old"] or "")[:120]
        new = (ch["new"] or "")[:120]
        if old != new:
            lines.append(f"    {label}: 「{old}」 → 「{new}」")
    return lines


def main() -> None:
    new_data = load_xlsx(XLSX)
    old_data = json.loads(OLD.read_text(encoding="utf-8"))

    report: dict = {"summary": {}, "sheet_diffs": {}, "code_gaps": {}}
    txt: list[str] = ["# Grimhand v0.8 总览表 vs 仓库快照 差异报告", ""]

    all_sheets = sorted(set(new_data) | set(old_data))
    for sheet in all_sheets:
        diff = diff_sheet(sheet, new_data.get(sheet, []), old_data.get(sheet, []))
        report["sheet_diffs"][sheet] = {
            "new_row_count": len(new_data.get(sheet, [])),
            "old_row_count": len(old_data.get(sheet, [])),
            **diff,
        }
        report["summary"][sheet] = {
            "added": len(diff["added_rows"]),
            "removed": len(diff["removed_rows"]),
            "changed": len(diff["changed_rows"]),
        }

    relic_db = (ROOT / "Scripts" / "Expedition" / "RelicDatabase.cs").read_text(encoding="utf-8")
    code_relics = parse_relic_ids(relic_db)
    excel_relics: dict[str, list] = {}
    for row in new_data.get("遗物", [])[2:]:
        if row and row[0] and row[0] != "遗物ID":
            excel_relics[norm(row[0])] = row

    report["code_gaps"]["relics_excel_not_code"] = sorted(set(excel_relics) - code_relics)
    report["code_gaps"]["relics_code_not_excel"] = sorted(code_relics - set(excel_relics))

    assets = index_card_assets()
    excel_cards: dict[str, list] = {}
    for row in new_data.get("卡牌", [])[1:]:
        if len(row) > 1 and row[1]:
            excel_cards[norm(row[1])] = row
    report["code_gaps"]["cards_excel_no_asset"] = sorted(set(excel_cards) - set(assets))

    OUT_JSON.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    txt.append("## 一、各 Sheet 变更统计（相对仓库 `_v08_excel_full.json`）")
    txt.append("")
    for sheet in all_sheets:
        s = report["summary"][sheet]
        if s["added"] or s["removed"] or s["changed"]:
            txt.append(
                f"- **{sheet}**：新增 {s['added']}，删除 {s['removed']}，修改 {s['changed']}"
                f"（行数 {report['sheet_diffs'][sheet]['old_row_count']} → {report['sheet_diffs'][sheet]['new_row_count']}）"
            )
    txt.append("")

    for sheet in all_sheets:
        diff = report["sheet_diffs"][sheet]
        if not (diff["added_rows"] or diff["removed_rows"] or diff["changed_rows"]):
            continue
        txt.append(f"## 二、{sheet}")
        if diff["added_rows"]:
            txt.append("")
            txt.append("### 新增")
            for item in diff["added_rows"]:
                row = item["row"]
                preview = " | ".join(x for x in row[:8] if x)
                txt.append(f"- **{item['key']}**：{preview}")
        if diff["removed_rows"]:
            txt.append("")
            txt.append("### 删除")
            for item in diff["removed_rows"]:
                preview = " | ".join(x for x in item["row"][:8] if x)
                txt.append(f"- **{item['key']}**：{preview}")
        if diff["changed_rows"]:
            txt.append("")
            txt.append("### 修改")
            for item in diff["changed_rows"]:
                txt.extend(format_change(sheet, item))
        txt.append("")

    txt.append("## 三、Excel vs 当前代码缺口")
    txt.append("")
    if report["code_gaps"]["relics_excel_not_code"]:
        txt.append("### 遗物：Excel 有、代码未实现")
        for rid in report["code_gaps"]["relics_excel_not_code"]:
            row = excel_relics[rid]
            name = row[1] if len(row) > 1 else ""
            desc = row[4] if len(row) > 4 else ""
            txt.append(f"- `{rid}` {name}：{desc[:100]}")
    if report["code_gaps"]["relics_code_not_excel"]:
        txt.append("")
        txt.append("### 遗物：代码有、Excel 已删/未列")
        for rid in report["code_gaps"]["relics_code_not_excel"]:
            txt.append(f"- `{rid}`")
    if report["code_gaps"]["cards_excel_no_asset"]:
        txt.append("")
        txt.append(f"### 卡牌：Excel 有、无对应 Card_*.asset（共 {len(report['code_gaps']['cards_excel_no_asset'])} 张）")
        for name in report["code_gaps"]["cards_excel_no_asset"][:50]:
            txt.append(f"- {name}")
        if len(report["code_gaps"]["cards_excel_no_asset"]) > 50:
            txt.append(f"- … 另有 {len(report['code_gaps']['cards_excel_no_asset']) - 50} 张")

    OUT_TXT.write_text("\n".join(txt), encoding="utf-8")
    print(f"Wrote {OUT_JSON}")
    print(f"Wrote {OUT_TXT}")
    for sheet, s in report["summary"].items():
        if s["added"] or s["removed"] or s["changed"]:
            print(f"{sheet}: +{s['added']} -{s['removed']} ~{s['changed']}")


if __name__ == "__main__":
    main()
