"""Dump Grimhand v0.9 overview Excel to text for inspection."""
import openpyxl
from pathlib import Path

XLSX = Path(r"C:\Users\Kelthuzad\Desktop\Grimhand实际内容总览表v0.9.xlsx")
OUT_DIR = Path(__file__).parent / "_v09_excel_dump"
OUT_DIR.mkdir(exist_ok=True)

wb = openpyxl.load_workbook(XLSX, data_only=True)
print("Sheets:", wb.sheetnames)

for name in wb.sheetnames:
    ws = wb[name]
    out = OUT_DIR / f"{name}.txt"
    with out.open("w", encoding="utf-8") as f:
        f.write(f"=== Sheet: {name} (dims={ws.dimensions}, rows={ws.max_row}, cols={ws.max_column}) ===\n")
        for row in ws.iter_rows(values_only=True):
            cells = ["" if v is None else str(v) for v in row]
            # skip fully empty rows
            if any(c.strip() for c in cells):
                f.write(" | ".join(cells) + "\n")
    print(f"Wrote {out}")
