#!/usr/bin/env python3
"""Dump every sheet of v0.81 overview Excel to a UTF-8 text file for review."""
from __future__ import annotations

import os
import sys
from pathlib import Path

import openpyxl

SRC = Path(r"c:\Users\Kelthuzad\Desktop\Grimhand实际内容总览表v0.81.xlsx")
OUT_DIR = Path(__file__).resolve().parent / "_v081_dump"
OUT_DIR.mkdir(exist_ok=True)

wb = openpyxl.load_workbook(SRC, data_only=True)

summary_lines = []
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    safe = "".join(c if c.isalnum() or c in "-_." else "_" for c in sheet_name)
    out_path = OUT_DIR / f"{safe}.txt"
    lines = [f"=== Sheet: {sheet_name}  (dims={ws.dimensions}) ==="]
    for row in ws.iter_rows(values_only=True):
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in row):
            continue
        cells = [("" if v is None else str(v)) for v in row]
        lines.append("\t".join(cells))
    out_path.write_text("\n".join(lines), encoding="utf-8")
    summary_lines.append(f"{sheet_name}\t{ws.max_row}x{ws.max_column}\t{out_path.name}")

summary = "\n".join(summary_lines)
(OUT_DIR / "_index.txt").write_text(summary, encoding="utf-8")
sys.stdout.buffer.write(("DUMPED:\n" + summary + "\n").encode("utf-8"))
