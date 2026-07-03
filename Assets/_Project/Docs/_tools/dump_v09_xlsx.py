# -*- coding: utf-8 -*-
import json
import openpyxl
from pathlib import Path

path = Path(r"c:\Users\Kelthuzad\Desktop\Grimhand实际内容总览表v0.9.xlsx")
out = Path(r"c:\Users\Kelthuzad\Documents\GitHub\Grimhand\Grimhand\Assets\_Project\Docs\_tools\_v09_xlsx_dump.json")

wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
data = {"sheets": wb.sheetnames, "content": {}}
for name in wb.sheetnames:
    ws = wb[name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(["" if c is None else str(c).strip() for c in row])
    data["content"][name] = rows

out.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
print("written", out, "sheets", len(wb.sheetnames))
