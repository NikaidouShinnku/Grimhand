"""生成远征事件参考 Excel（不含祭坛）。"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parents[1] / "ExpeditionEvents_Reference.xlsx"

EVENTS = [
    {
        "id": "mysterious_traveler",
        "name": "神秘旅者",
        "scene": "戴兜帽的旅者摊开手掌，展示发光的物品。\n「交易，还是离开？」",
        "choices": [
            ("A", "用 30 金币购买", "消耗 30 金", "随机卡牌奖励（蓝优先，否则白）", "奖励领取界面", "已实现"),
            ("B", "接受礼物", "—", "随机遗物 + 随机队员牌组加入诅咒牌「混沌之触」", "消息提示 + 遗物领取界面", "已实现"),
            ("C", "拒绝离开", "—", "无事发生", "—", "已实现"),
        ],
    },
    {
        "id": "ancient_temple",
        "name": "古老神殿",
        "scene": "残破神殿中祭台火焰仍在燃烧，神像似乎在注视着你。",
        "choices": [
            ("A", "虔诚祈祷", "全队 -10% HP", "远征期间全队 ATK+1", "扣血动画", "已实现"),
            ("B", "亵渎圣堂", "—", "50 金币；下场战斗敌人 ATK+20%", "金币领取界面", "已实现"),
            ("C", "静默离开", "—", "无事发生", "—", "已实现"),
        ],
    },
    {
        "id": "injured_adventurer",
        "name": "受伤的冒险者",
        "scene": "倒地的冒险者仍在流血：「求你…帮帮我…」",
        "choices": [
            ("A", "救治", "全队 -15% HP", "随机遗物", "扣血动画 → 遗物领取", "已实现"),
            ("B", "搜刮", "—", "20 金币 + 随机卡牌；触发「冒险者复仇」前置", "奖励领取界面", "已实现"),
            ("C", "无视", "—", "无事发生", "—", "已实现"),
        ],
    },
    {
        "id": "magic_spring",
        "name": "魔法泉水",
        "scene": "荧光泉水映照出奇异影像。",
        "choices": [
            ("A", "饮用泉水", "随机", "60% 全队 +25% HP / 25% 1 人 ATK+2 / 15% 全队 -15% HP", "治疗或扣血/选人动画", "已实现"),
            ("B", "装瓶带走", "—", "获得 2×泉水瓶", "—", "已实现"),
            ("C", "不碰", "—", "无事发生", "—", "已实现"),
        ],
    },
    {
        "id": "gambler_dice",
        "name": "赌徒的骰子",
        "scene": "矮人转着发光骰子：「来玩一把？」",
        "choices": [
            ("A", "小赌（20 金币）", "20 金", "50% +50 金", "金币领取界面", "已实现"),
            ("B", "大赌（全部金币）", "全部金币", "40% 翻倍 / 30% 清零 / 30% 稀有遗物", "金币或遗物领取", "已实现"),
            ("C", "不赌", "—", "无事发生", "—", "已实现"),
        ],
    },
    {
        "id": "mirror_phantom",
        "name": "镜中幻影",
        "scene": "魔法镜中映出会动的队伍影子。",
        "choices": [
            ("A", "进入镜中挑战", "—", "镜像战斗；胜利获得蓝色卡牌", "战斗 → 卡牌领取", "已实现"),
            ("B", "打碎镜子", "—", "镜之碎片 ×1", "—", "已实现"),
            ("C", "离开", "—", "无事发生", "—", "已实现"),
        ],
    },
    {
        "id": "cursed_bookshelf",
        "name": "被诅咒的书架",
        "scene": "一本书在自行翻页，文字不断变化。",
        "choices": [
            ("A", "阅读", "随机 1 人 -10 HP", "随机蓝色卡牌", "单人扣血 → 卡牌领取", "已实现"),
            ("B", "撕页带走", "—", "古卷残页 ×1", "—", "已实现"),
            ("C", "合上书", "—", "无事发生", "—", "已实现"),
        ],
    },
    {
        "id": "adventurer_revenge",
        "name": "冒险者的复仇",
        "scene": "被你搜刮过的冒险者带着同伴出现了。",
        "prerequisite": "需先搜刮受伤冒险者",
        "choices": [
            ("A", "道歉赔偿（40 金币）", "40 金", "和解；下 3 层节点类型可见", "—", "已实现"),
            ("B", "应战", "—", "2 骷髅兵战斗；胜利 +30 金", "战斗 → 金币领取", "已实现"),
            ("C", "逃跑", "全队 -5% HP", "—", "扣血动画", "已实现"),
        ],
    },
    {
        "id": "training_dummy",
        "name": "训练人偶",
        "scene": "破旧训练人偶仍可用于练习。",
        "choices": [
            ("A", "全队训练", "全队 -10% HP", "远征 DEF+1", "扣血动画", "已实现"),
            ("B", "单人特训", "1 人 -20% HP", "该角色 ATK+2", "选人扣血 → 同角色加攻", "已实现"),
            ("C", "休息", "—", "全队 +10% HP", "治疗动画", "已实现"),
        ],
    },
    {
        "id": "soul_rift",
        "name": "灵魂裂隙",
        "scene": "紫色能量从空间裂缝中涌出。",
        "choices": [
            ("A", "吸收能量", "—", "能量上限 +1；每场战斗开始随机 1 人 -5 HP", "战斗开始时扣血", "已实现"),
            ("B", "封印裂隙", "移除 1 张卡牌", "随机稀有遗物", "选卡面板 → 遗物领取", "已实现"),
            ("C", "绕行", "—", "无事发生", "—", "已实现"),
        ],
    },
    {
        "id": "wandering_smith",
        "name": "流浪铁匠",
        "scene": "驼背铁匠的炉火仍在燃烧。",
        "choices": [
            ("A", "强化卡牌（15 金币）", "15 金", "1 张卡牌升 1 级", "背包式选卡 + 确认", "已实现"),
            ("B", "融合卡牌", "销毁 2 张同类型牌", "获得更高品质牌", "两步选卡 + 确认", "已实现"),
            ("C", "离开", "—", "无事发生", "—", "已实现"),
        ],
    },
    {
        "id": "tired_camp",
        "name": "疲惫营地",
        "scene": "废弃营地余烬未熄，可以休整。",
        "choices": [
            ("A", "深度休息", "—", "跳过下一层路线；全队 +30% HP", "治疗动画", "已实现"),
            ("B", "简单休息", "—", "全队 +15% HP", "治疗动画", "已实现"),
            ("C", "搜索营地", "—", "10–25 随机金币", "金币领取界面", "已实现"),
        ],
    },
    {
        "id": "jade_workshop",
        "name": "玉匠工坊",
        "scene": "老工匠看到你的翡翠原石，眼睛一亮。",
        "prerequisite": "需持有翡翠原石",
        "choices": [
            ("A", "打磨为戒指", "—", "翡翠原石 → 翡翠戒指", "—", "已实现"),
            ("B", "雕刻为短刀", "—", "翡翠原石 → 翡翠短刀", "—", "已实现"),
            ("C", "离开", "—", "无事发生", "—", "已实现"),
        ],
    },
    {
        "id": "ancient_furnace",
        "name": "古老熔炉",
        "scene": "远古熔炉仍在燃烧，靴子似乎有所回应。",
        "prerequisite": "需持有燃烬之靴",
        "choices": [
            ("A", "以血淬火", "全队 -10% HP", "燃烬之靴 → 赤红烈焰靴", "扣血动画", "已实现"),
            ("B", "保留原样", "—", "无事发生", "—", "已实现"),
        ],
    },
    {
        "id": "abyss_whisper",
        "name": "深渊低语",
        "scene": "黑暗中的呢喃让你感到诱惑。",
        "prerequisite": "队伍需有恶魔（游侠）",
        "choices": [
            ("A", "倾听低语", "恶魔 -20% HP", "恶魔获得专属紫卡", "单人扣血 → 卡牌领取", "已实现"),
            ("B", "献出记忆", "移除 1 张卡牌", "全队 ATK+1", "选卡面板", "已实现"),
            ("C", "离开", "—", "无事发生", "—", "已实现"),
        ],
    },
]

HEADERS = [
    "事件 ID",
    "事件名称",
    "场景描述",
    "前置条件",
    "选项",
    "选项标题",
    "代价",
    "效果",
    "视觉反馈",
    "实现状态",
]


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "远征事件"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    for col, title in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap

    row = 2
    for evt in EVENTS:
        prereq = evt.get("prerequisite", "")
        first = True
        for choice in evt["choices"]:
            ws.cell(row=row, column=1, value=evt["id"] if first else "")
            ws.cell(row=row, column=2, value=evt["name"] if first else "")
            ws.cell(row=row, column=3, value=evt["scene"] if first else "")
            ws.cell(row=row, column=4, value=prereq if first else "")
            ws.cell(row=row, column=5, value=choice[0])
            ws.cell(row=row, column=6, value=choice[1])
            ws.cell(row=row, column=7, value=choice[2])
            ws.cell(row=row, column=8, value=choice[3])
            ws.cell(row=row, column=9, value=choice[4])
            ws.cell(row=row, column=10, value=choice[5])
            for col in range(1, 11):
                ws.cell(row=row, column=col).alignment = wrap
            first = False
            row += 1

    widths = [22, 14, 36, 18, 6, 22, 16, 36, 22, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
