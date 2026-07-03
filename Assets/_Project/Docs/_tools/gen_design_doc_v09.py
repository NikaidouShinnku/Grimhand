# -*- coding: utf-8 -*-
"""从 Grimhand实际内容总览表v0.9.xlsx 生成策划案 v0.9（Markdown）。"""
from __future__ import annotations

import re
from datetime import date
from pathlib import Path

import openpyxl

XLSX = Path(r"c:\Users\Kelthuzad\Desktop\Grimhand实际内容总览表v0.9.xlsx")
OUT = Path(r"c:\Users\Kelthuzad\Desktop\Grimhand_Design_Doc_v0.9.md")

CHARS = ["战士", "法老", "恶魔", "毒蛇女王", "巫妖女王"]


def cell(row, idx):
    if idx >= len(row):
        return ""
    v = row[idx]
    return "" if v is None else str(v).strip()


def md_table(headers, rows):
    if not rows:
        return "_（无数据）_\n"
    lines = ["| " + " | ".join(headers) + " |", "| " + " | ".join(["---"] * len(headers)) + " |"]
    for r in rows:
        lines.append("| " + " | ".join(str(c).replace("\n", "<br>") for c in r) + " |")
    return "\n".join(lines) + "\n"


def load_sheet(wb, name):
    ws = wb[name]
    return [list(r) for r in ws.iter_rows(values_only=True)]


def parse_notes_sheet(rows):
    """说明（经验值和关键词等）"""
    keywords = []
    card_colors = []
    in_level_rows = []
    hand_rows = []
    energy_rows = []
    meta_rows = []

    for row in rows[1:]:
        r = [cell(row, i) for i in range(max(len(row), 18))]
        kw, kw_desc = r[13], r[14]
        color, color_desc = r[16], r[17]
        if kw:
            keywords.append((kw, kw_desc))
        if color and color_desc:
            card_colors.append((color, color_desc))
        if r[0].isdigit():
            in_level_rows.append((r[0], r[1]))
        if r[3].isdigit() or r[3] in ("5", "6", "7", "8"):
            if r[3] and r[4]:
                hand_rows.append((r[3], r[4]))
        if r[6].isdigit() or r[6] in ("8", "9", "10"):
            if r[6] and r[7]:
                energy_rows.append((r[6], r[7]))
        if r[9].isdigit():
            meta_rows.append((r[9], r[10], r[11]))

    timing_notes = []
    for row in rows:
        r = [cell(row, i) for i in range(18)]
        if r[13].startswith("【回合") or r[13] in ("【快速启动】", "【继承】"):
            timing_notes.append((r[13], r[14]))

    return {
        "keywords": keywords,
        "card_colors": card_colors,
        "in_level": in_level_rows,
        "hand": hand_rows,
        "energy": energy_rows,
        "meta": meta_rows,
        "timing": timing_notes,
    }


def parse_characters_talents(rows):
    stats = {c: [] for c in CHARS}
    # col offsets per character in stat block
    stat_cols = {"战士": 1, "法老": 4, "恶魔": 7, "毒蛇女王": 10, "巫妖女王": 13}
    talent_cols = {"战士": (18, 19, 20), "法老": (21, 22, 23), "恶魔": (24, 25, 26),
                   "毒蛇女王": (27, 28, 29), "巫妖女王": (30, 31, 32)}

    for row in rows[3:]:
        r = [cell(row, i) for i in range(max(len(row), 33))]
        level = r[0]
        if not level or not level[0].isdigit():
            continue
        for char, base in stat_cols.items():
            hp = r[base]
            xp = r[base + 1]
            spd = r[base + 2]
            if hp:
                stats[char].append((level, hp, xp, spd))
        tc = talent_cols
        for char, (lv_i, s1_i, s2_i) in talent_cols.items():
            lv = r[lv_i]
            if not lv or not lv[0].isdigit():
                continue
            s1, s2 = r[s1_i], r[s2_i]
            if s1 or s2:
                if not any(t[0] == lv for t in stats.get("_talents_" + char, [])):
                    stats.setdefault("_talents_" + char, []).append((lv, s1, s2))
                else:
                    for i, t in enumerate(stats["_talents_" + char]):
                        if t[0] == lv:
                            stats["_talents_" + char][i] = (lv, s1 or t[1], s2 or t[2])

    talents = {c: stats.pop("_talents_" + c, []) for c in CHARS}
    return stats, talents


def is_test_card(card):
    """游戏内测试卡，不写入正式策划案。"""
    if card.get("kind") == "测试用":
        return True
    note = card.get("note") or ""
    return "测试卡" in note or "无法获取" in note


def parse_cards(rows):
    cards = []
    starters = []
    for row in rows[1:]:
        r = [cell(row, i) for i in range(max(len(row), 19))]
        if r[0] and r[1] and r[0] != "角色":
            cards.append({
                "char": r[0], "name": r[1], "cost": r[2], "type": r[3], "kind": r[4],
                "effect": r[5], "rarity": r[6], "upgrades": r[7], "upgrade_eff": r[8],
                "upgrade_xp": r[9], "note": r[10],
            })
        if len(r) > 18 and r[16] and r[17]:
            starters.append((r[16], r[17], r[18]))
    return cards, starters


def parse_table_sheet(rows, header_row=1, min_cols=3):
    if len(rows) <= header_row:
        return [], []
    header = [cell(rows[header_row], i) for i in range(len(rows[header_row]))]
    while header and not header[-1]:
        header.pop()
    body = []
    for row in rows[header_row + 1:]:
        r = [cell(row, i) for i in range(len(header))]
        if any(r):
            body.append(r)
    return header, body


def section_meta_progression():
    return """### 2.0 经验与祭坛升级（v0.9 正式规则）

- **经验来源：** 战斗胜利、特殊事件等途径获得 **共享经验池**。
- **经验用途：** 在 **祭坛** 消耗经验，升级远征层面的各项能力（见下表）。
- **可升级项：**
  - 每回合 **抽牌数量**（§2.2；总览表表头写作「手牌数升级表」）
  - **能量上限**（§2.3）
  - 单张 **卡牌强化**（§5.1「升级所需经验」，如 8/级、15/级）
  - 角色 **局外等级** → 解锁天赋节点（§2.4、§4.2）
- **手牌上限：** 固定 **10**，不可升级。
- **局内等级：** 角色 HP/SPD 随局内等级成长（§4.1）；经验表见 §2.1。

"""


def section_notes(notes):
    lines = ["## 2. 全局数值与关键词（来自总览表·说明 sheet）\n"]
    lines.append(section_meta_progression())
    lines.append("### 2.1 局内等级经验（单角色战斗内升级）\n")
    lines.append(md_table(["局内等级", "升至该级所需XP"], notes["in_level"]))
    lines.append("### 2.2 每回合抽牌数升级（总览表表头：「手牌数升级表」）\n")
    lines.append(md_table(["每回合抽牌数", "所需XP"], notes["hand"]))
    lines.append("### 2.3 能量上限升级\n")
    lines.append(md_table(["能量上限", "所需XP"], notes["energy"]))
    lines.append("### 2.4 局外等级经验（角色 meta 成长解锁天赋）\n")
    lines.append(md_table(["局外等级", "本级需XP", "累计XP"], notes["meta"]))
    lines.append("### 2.5 关键词\n")
    lines.append(md_table(["关键词", "描述"], patch_keyword_burn(notes["keywords"])))
    lines.append("### 2.6 卡牌稀有度颜色\n")
    lines.append(md_table(["颜色", "对应稀有度"], notes["card_colors"]))
    lines.append("### 2.7 时序用语\n")
    lines.append(md_table(["用语", "定义"], notes["timing"]))
    return "\n".join(lines)


def patch_keyword_burn(keywords):
    out = []
    for k, d in keywords:
        if "灼烧" in k:
            d = "回合开始每层造成2伤害"
        out.append((k, d))
    return out


def section_turn_lifecycle():
    return """### 3.2 完整回合循环

一场战斗内，每个 **回合 N** 按以下顺序循环（玩家视角：抽牌 → 选牌确认 → 看演出 → 下一回合）：

| 序 | 阶段 | 说明 |
|----|------|------|
| ① | **回合末** | 弃手、状态持续递减、护甲清零（例外见卡牌）；逻辑在 **上次演出播完后** 执行 |
| ② | **抽牌** | 回能 → **回合开始** 状态 tick（中毒、**灼烧** 等）→ 双方抽牌 |
| ③ | **规划** | 敌方 AI 选牌并展示意图（部分隐藏）；玩家选牌/目标/扣费；可 **快速启动** 立即出牌 |
| ④ | **速度结算** | 敌我卡牌按 **速度轮询** 交错出牌（非「先玩家后敌人」）；应对卡可插队 |
| ⑤ | **演出** | ④ 逻辑一次性算完，动画逐段播放；播完后才进入 ① |

**敌人回合：** 不存在单独的「敌人回合阶段」。敌牌与玩家牌在同一 **速度 schedule** 中按 round+速度 交错结算。

**非战斗时段（灵能体）：** 从 **④ 速度结算逻辑结束** 到 **下次玩家点击确认出牌** 之间的全部时间，包括 ① 回合末、② 回合开始 tick、③ 规划期。**不包含** 速度阶段内任意卡牌效果。此期间造成的伤害（含中毒/灼烧 tick、延迟伤害）受 **灵能体 +20%** 加成。

"""


def section_confirmed_mechanics():
    return """### 3.6 已定稿特殊机制

| 卡牌/主题 | 规则 |
|-----------|------|
| **灵能体** | 非战斗时段造成的伤害 +20%（见 §3.2） |
| **战术大师的终结技** | 伤害 = **整场远征**（从进入远征到结束）累计 **应对成功次数** × 5，跨战斗场不清零 |
| **灵界封印** | 敌方 **下一张** 牌：进入弃牌堆、**不消耗** 敌方能量、**效果不生效**；本卡 X 费 = 被封印牌费用 + 1 |
| **阿努比斯化身** | **本场战斗临时** +50% 最大 HP，**当前 HP 同比填充**（例 80/80→120/120，40/80→80/120）；+50% 增伤与强固；接下来 2 回合不能出牌；战斗结束恢复 |
| **铁壁弹反（演出）** | 敌攻动画 **完整播完后**，战士以 **Attack 立绘** 段造成反射伤害 |
| **灼烧** | **回合开始** tick（与中毒同窗口），每层 2 伤 |

"""


def section_battle_framework(notes):
    kw_map = {k: d for k, d in patch_keyword_burn(notes["keywords"])}
    lines = [
        "## 3. 战斗系统（v0.9 框架，继承 v0.8 结构 + 总览表定义）\n",
        "### 3.1 核心原则\n",
        "- 角色只有 **HP** 与 **SPD**，已删除 ATK/DEF（总览表·角色 sheet 标题注明 v0.7 已删）。",
        "- 卡牌伤害/护甲/治疗为 **固定数值**；变强来自升级、遗物、天赋、临时 Buff。",
        "- 站位 **无隐式输出/承伤倍率**；差异通过卡牌【前/中/后】、Reach、嘲讽、天赋体现。",
        "- 承伤手段：护甲（回合末清零，除非卡牌/天赋例外）、减伤%、闪避、虚化、守护/嘲讽等。\n",
    ]
    lines.append(section_turn_lifecycle())
    lines.append("### 3.3 能量与牌堆（基础值，可被升级/遗物修改）\n")
    lines.extend([
        "| 项 | 基础值 | 升级路径（见 §2.3 / §2.2） |",
        "|---|---|---|",
        "| 能量上限 | 8 | 9（100XP）、10（200XP） |",
        "| 首回合能量 | 回满至上限 | — |",
        "| 之后每回合 | +4，不超过上限 | — |",
        "| 每回合抽牌 | 5（见 §2.2 祭坛升级） | 6（50XP）、7（100XP）、8（200XP） |",
        "| 手牌上限 | **10**（固定） | 不可升级 |\n",
        "### 3.4 应对机制\n",
    ])
    for key in ["【应对攻击】", "【应对状态】", "【应对防御】"]:
        if key in kw_map:
            lines.append(f"- **{key}**：{kw_map[key]}")
    lines.append(
        "\n**效果描述标点规则（v0.9 正式）：**\n"
        "- 以【应对攻击】/【应对状态】/【应对防御】开头的卡牌：**第一个句号之前** "
        "（从应对关键词起，含句内逗号分隔的各段）= **必须应对成功** 才触发。\n"
        "- **句号之后** 的内容 = 只要这张卡 **在本回合生效**（无论因应对插队，"
        "还是未应对到、按速度轮到正常结算），该段 **必定触发**。\n"
        "- 「若应对失败…」等明示失败分支 = 独立第三段，按文案字面执行。\n"
        "- **卡牌/怪物效果以总览表「效果」列原文为准**，不做额外推断或改写。\n"
    )
    lines.append("### 3.5 状态效果 tick\n")
    for key in ["【中毒×层数】", "【灼烧×层数】", "【减速×层数】", "【加速×层数】", "【虚化】", "【闪避】"]:
        if key in kw_map:
            lines.append(f"- **{key}**：{kw_map[key]}")
    lines.append("- **【灼烧×层数】** 与中毒均在 **回合开始**（抽牌阶段）tick。\n")
    lines.append(section_confirmed_mechanics())
    lines.append("### 3.7 战斗演出要点\n")
    lines.append(
        "- **速度结算**：Commit 后逻辑一次算完；立绘按 **PortraitPoseChanged → 效果 → PortraitIdleRestored** 分段播放。\n"
        "- **玩家/敌人普通牌**：有移中、Attack/Defense 立绘、归位。\n"
        "- **应对步（除弹反反击段）**：通常 **无** 出牌立绘；减伤/免疫体现在 **敌攻段** 的 Defense 受击。\n"
        "- **铁壁弹反**：敌攻段结束后，**独立段** 播放应对者 Attack 立绘造成反射伤害。\n"
        "- **状态 tick**（中毒/灼烧）：独立段，无出牌立绘；脚边状态图标在 StatusApplied 动画后更新。\n"
        "- **快速启动**：规划阶段立即播放，走标准出牌段，不进速度队列。\n"
    )
    lines.append("")
    return "\n".join(lines)


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    parts = []

    parts.append(f"""# GRIMHAND — 游戏策划案 v0.9

**Game Design Document v0.9**  
**日期：** {date.today().isoformat()}  
**状态：** 由《Grimhand实际内容总览表v0.9.xlsx》导出；取代 v0.8 为当前内容权威  
**来源文件：** `{XLSX.name}`

> **权威原则：** 卡牌、怪物、遗物、事件的 **效果描述以总览表原文为准**，本文档忠实导出，不做臆测解读。  
> 与当前代码实现不一致处 **以本文档 / 总览表为准**；实现偏差见 `战斗逻辑及机制参考.docx`。

---

## 1. 版本变更摘要（相对 v0.8）

| 变更 | 说明 |
|------|------|
| 新增可玩角色 | 新增 **毒蛇女王**、**巫妖女王**（五角色） |
| 远征编队 | 固定 **3 人** |
| 关键词 | 以 v0.9 总览表为准；【增伤】【减伤】【易伤】【虚弱】等均为 **±1%/层** 百分比体系 |
| 灼烧 / 减速 | 灼烧：**回合开始** 每层 2 伤；减速：每层 **-1 SPD** |
| 升级体系 | 战斗/事件获 XP → **祭坛** 升级抽牌数、能量上限、卡牌强化、局外天赋；**手牌上限固定 10** |
| 卡牌库 | 全量卡牌、初始牌组、怪物、Boss、遗物、事件以 xlsx 为准（不含游戏内测试卡） |

---

""")

    notes_rows = load_sheet(wb, "说明（经验值和关键词等）")
    notes = parse_notes_sheet(notes_rows)
    parts.append(section_notes(notes))
    parts.append(section_battle_framework(notes))

    # Characters
    char_rows = load_sheet(wb, "角色和天赋")
    stats, talents = parse_characters_talents(char_rows)
    parts.append("\n## 4. 角色与天赋\n")
    parts.append("### 4.1 局内属性成长（HP / SPD）\n")
    for char in CHARS:
        parts.append(f"#### {char}\n")
        parts.append(md_table(["局内等级", "HP", "升级XP", "SPD"], stats[char]))
    parts.append("### 4.2 局外天赋树\n")
    parts.append(
        "规则：每角色 **两个天赋槽位**；局外等级提升时解锁节点，玩家选用节点增强战斗。"
        "描述未指明目标时 **仅影响自身**。\n"
    )
    for char in CHARS:
        parts.append(f"#### {char}\n")
        parts.append(md_table(["局外等级", "槽位1", "槽位2"], talents[char]))

    # Cards
    card_rows = load_sheet(wb, "卡牌")
    cards, starters = parse_cards(card_rows)
    parts.append("\n## 5. 卡牌系统\n")
    parts.append("### 5.1 字段说明\n")
    parts.append("| 字段 | 含义 |\n|------|------|\n")
    parts.append("| 职业 | 职业卡 / 通用 / 测试用 |\n")
    parts.append("| 可升级次数 | 祭坛卡牌强化上限 |\n")
    parts.append("| 每次升级效果 | 单级数值变化 |\n")
    parts.append("| 升级所需经验 | 祭坛消耗共享 XP，如 `8/级` 或 `15/级` |\n")
    parts.append("\n> 游戏内测试卡（如「作者境的一击」）不列入正式策划案。\n")
    parts.append("\n### 5.2 初始牌组\n")
    starter_by_char = {}
    for ch, name, count in starters:
        if not ch or not name:
            continue
        starter_by_char.setdefault(ch, []).append((name, count))
    for char in CHARS + ["通用"]:
        if char not in starter_by_char:
            continue
        parts.append(f"#### {char}\n")
        parts.append(md_table(["卡牌", "数量"], starter_by_char[char]))

    parts.append("\n### 5.3 全卡牌列表（按角色）\n")
    cards = [c for c in cards if not is_test_card(c)]
    for char in CHARS + ["通用", "诅咒", "消耗品"]:
        subset = [c for c in cards if c["char"] == char]
        if not subset:
            continue
        parts.append(f"#### {char}\n")
        rows = [(c["name"], c["cost"], c["type"], c["rarity"], c["effect"], c["upgrades"], c["upgrade_eff"], c["upgrade_xp"], c["note"]) for c in subset]
        parts.append(md_table(["名称", "费", "类型", "稀有", "效果", "可升级", "升级效果", "升级XP", "说明"], rows))

    # catch-all other characters in sheet
    known = set(CHARS + ["通用", "诅咒", "消耗品"])
    other_chars = sorted({c["char"] for c in cards if c["char"] not in known})
    for char in other_chars:
        subset = [c for c in cards if c["char"] == char]
        parts.append(f"#### {char}\n")
        rows = [(c["name"], c["cost"], c["type"], c["rarity"], c["effect"]) for c in subset]
        parts.append(md_table(["名称", "费", "类型", "稀有", "效果"], rows))

    # Expedition
    parts.append("\n## 6. 远征结构\n")
    parts.append(
        "- **编队人数：** 固定 **3 名** 可玩角色。\n"
        "- **经验：** 战斗胜利、特殊事件等获得共享 XP，于祭坛分配（§2.0）。\n"
        "- **内容池：** 卡牌获取、遗物、事件、怪物组合见对应章节；数值以总览表为准。\n"
    )

    # Relics
    parts.append("\n## 7. 遗物\n")
    rel_rows = load_sheet(wb, "遗物")
    parts.append(f"{cell(rel_rows[0], 0)}\n\n")
    parts.append("> 部分遗物「每 20 层成长」的后续数值总览表尚未补全，以当前表内列为准；层数体系后续可能调整。\n\n")
    h, body = parse_table_sheet(rel_rows, 1)
    parts.append(md_table(h[:6], [r[:6] for r in body if r[0] and r[0] != "遗物ID"]))

    # Consumables - if sheet exists
    if "消耗品" in wb.sheetnames:
        parts.append("\n## 8. 消耗品\n")
        con_rows = load_sheet(wb, "消耗品")
        h, body = parse_table_sheet(con_rows, 0)
        if not h or h[0] in ("", "消耗品"):
            h, body = parse_table_sheet(con_rows, 1)
        parts.append(md_table(h[:6] if h else ["列1"], [r[:6] for r in body if any(r)]))

    # Boss
    parts.append("\n## 9. Boss 设计\n")
    boss_rows = load_sheet(wb, "Boss设计")
    for i, row in enumerate(boss_rows[:3]):
        t = cell(row, 0)
        if t:
            parts.append(f"> {t}\n")
    # dump boss content as subsections
    buf = []
    current = ""
    for row in boss_rows[3:]:
        r = [cell(row, i) for i in range(12)]
        if not any(r):
            continue
        title = r[0]
        if title in ("Boss卡牌", "Boss基本数值") or title.endswith("王") or title.endswith("后"):
            if buf and current:
                parts.append(f"### {current}\n")
                parts.append("\n".join(buf) + "\n")
                buf = []
            if title not in ("Boss卡牌", "Boss基本数值"):
                current = title
            else:
                buf.append("| " + " | ".join(x for x in r if x) + " |")
        else:
            if r[1]:  # card row
                buf.append("| " + " | ".join(r[1:8]) + " |")
            elif r[0]:
                buf.append(f"- {r[0]}: " + ", ".join(x for x in r[1:8] if x))
    if buf and current:
        parts.append(f"### {current}\n")
        parts.append("\n".join(buf) + "\n")

    # Simpler boss dump - full table export
    parts.append("\n### 9.1 Boss 原始表（完整导出）\n")
    parts.append("```\n")
    for row in boss_rows:
        line = "\t".join(cell(row, i) for i in range(min(12, len(row) or 12)) if cell(row, i))
        if line.strip():
            parts.append(line + "\n")
    parts.append("```\n")

    # Minions
    parts.append("\n## 10. 小怪设计\n")
    min_rows = load_sheet(wb, "小怪设计")
    for row in min_rows[:5]:
        t = cell(row, 0)
        if t:
            parts.append(f"> {t}\n")
    parts.append("\n### 10.1 小怪原始表（完整导出）\n```\n")
    for row in min_rows[5:]:
        line = "\t".join(cell(row, i) for i in range(8) if cell(row, i))
        if line.strip():
            parts.append(line + "\n")
    parts.append("```\n")

    # Encounters
    if "怪物组合" in wb.sheetnames:
        parts.append("\n## 11. 怪物组合\n```\n")
        for row in load_sheet(wb, "怪物组合"):
            line = "\t".join(cell(row, i) for i in range(10) if cell(row, i))
            if line.strip():
                parts.append(line + "\n")
        parts.append("```\n")

    # Events
    parts.append("\n## 12. 远征事件\n")
    ev_rows = load_sheet(wb, "事件")
    parts.append("```\n")
    for row in ev_rows:
        line = "\t".join(cell(row, i) for i in range(8) if cell(row, i))
        if line.strip():
            parts.append(line + "\n")
    parts.append("```\n")

    # Icons
    if "icon设计" in wb.sheetnames:
        parts.append("\n## 13. Icon 设计\n```\n")
        for row in load_sheet(wb, "icon设计"):
            line = "\t".join(cell(row, i) for i in range(6) if cell(row, i))
            if line.strip():
                parts.append(line + "\n")
        parts.append("```\n")

    parts.append("""
---

## 14. 文档维护

- **权威来源：** `Grimhand实际内容总览表v0.9.xlsx`
- **导出脚本：** `Assets/_Project/Docs/_tools/gen_design_doc_v09.py`（改 xlsx 后重跑可刷新本文档）
- **战斗实现参考（代码现状）：** `Assets/_Project/Docs/战斗逻辑及机制参考.docx`

---

*本文档由总览表自动导出 + 框架章节组成。§5 卡牌、§9–§12 敌人/事件的效果列 **逐字以 xlsx 为准**；实现侧须对齐表内描述，不得自行改写语义（例：「无法使用攻击牌」≠「无法出牌」）。*
""")

    OUT.write_text("".join(parts), encoding="utf-8")
    print("OK", OUT, OUT.stat().st_size)


if __name__ == "__main__":
    main()
