# -*- coding: utf-8 -*-
"""从 Unity 数据导出怪物总览 Excel（格式对齐 Boss设计 / 小怪设计）。"""
import os
import re
import shutil
from collections import Counter, defaultdict
from copy import copy

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

PROJECT = r"c:\Users\Kelthuzad\Documents\GitHub\Grimhand\Grimhand"
REF_XLSX = r"c:\Users\Kelthuzad\Desktop\The Grimhands Asset\Grimhand实际内容总览表.xlsx"
OUT_XLSX = r"c:\Users\Kelthuzad\Desktop\Grimhand怪物总览表_2026-06-10.xlsx"

BOSS_IDS = {"char_skeleton_king", "char_ghost_queen", "char_explosive_skull"}

SLOT = {1: "前排", 2: "中排", 3: "后排"}
CARD_TYPE = {0: "攻击", 1: "防御", 2: "状态"}
RARITY = {0: "白", 1: "蓝", 2: "紫", 3: "橙"}
REACH = {0: "前/中/后", 1: "前/中", 2: "后排", 3: "中/后"}
TARGET = {
    0: "默认敌人", 1: "自身", 2: "友前排", 3: "友后排", 4: "上一行动者",
    5: "手动选择", 6: "敌前排", 7: "敌中排", 8: "敌后排", 9: "友前排槽",
    10: "友中排槽", 11: "友后排槽", 12: "全体敌人", 13: "随机敌人", 14: "随机N名敌人",
}
ACTION = {
    0: "伤害", 1: "护甲", 2: "治疗", 3: "状态", 4: "移除状态", 5: "换位",
    6: "下回合抽牌", 7: "抽牌", 8: "反射伤害", 9: "应对减伤",
    10: "阿努比斯化身", 11: "锁定出牌", 12: "减能量回复", 13: "伤害转嫁",
}
CONDITION = {0: "", 1: "【应对攻击】"}
KW_CN = {
    "melee": "近战", "far_shot": "远射", "snipe": "狙击", "aoe": "AOE",
    "guard": "防御", "slow": "减速", "poison": "中毒", "summon": "召唤",
    "parry": "应对攻击", "exhaust": "消耗", "self_destruct": "自毁",
    "bonus_hand": "额外手牌",
}
STATUS_CN = {
    "slow": "减速", "poison": "中毒", "bone_workshop": "骨之王座",
}
TRAIT_CN = {
    "boss_first_hit_block": "每回合Boss第一次受到伤害时，获得10护甲",
    "boss_turn_def_up": "每回合开始时永久+1基础防御",
    "skull_self_destruct_hand": "每回合必定使用骷髅自爆卡牌（不占抽牌数）",
    "ghost_queen_enrage": "HP首次低于120时虚化（攻击最多1伤）并下回合获得「幽灵女王之怒」",
}
PLAYER_ASSET_PREFIXES = ("Character_Knight", "Character_Mage", "Character_Ranger",
                         "Character_Pharaoh", "Character_Warrior", "Character_Demon", "Character_Devil")


def decode_unicode(s):
    if not s:
        return s
    try:
        return s.encode("utf-8").decode("unicode_escape")
    except Exception:
        return s


def read_text(path):
    with open(path, "r", encoding="utf-8") as f:
        return f.read()


def build_guid_map(folder):
    mapping = {}
    for name in os.listdir(folder):
        if not name.endswith(".meta"):
            continue
        text = read_text(os.path.join(folder, name))
        m = re.search(r"^guid:\s*(\w+)", text, re.M)
        if m:
            mapping[m.group(1)] = name.replace(".meta", "")
    return mapping


def parse_scalar(block, key, default=""):
    m = re.search(rf"^\s*{re.escape(key)}:\s*(.+)$", block, re.M)
    if not m:
        return default
    val = m.group(1).strip()
    if val.startswith('"') and val.endswith('"'):
        return decode_unicode(val.strip('"'))
    return val


def parse_int(block, key, default=0):
    val = parse_scalar(block, key, default)
    try:
        return int(val)
    except Exception:
        return default


def parse_list(block, key):
    lines = block.splitlines()
    items = []
    in_list = False
    key_re = re.compile(rf"^\s*{re.escape(key)}:\s*(.*)$")
    for line in lines:
        m = key_re.match(line)
        if m:
            rest = m.group(1).strip()
            if rest == "[]":
                return []
            in_list = True
            continue
        if in_list:
            if re.match(r"^\s+- ", line):
                items.append(line.strip()[2:].strip())
            elif re.match(r"^\s*[A-Za-z_]", line):
                break
    return items


def parse_guids(items):
    guids = []
    for item in items:
        m = re.search(r"guid:\s*(\w+)", item)
        if m:
            guids.append(m.group(1))
    return guids


def parse_actions(block):
    actions = []
    chunks = re.split(r"\n\s+- Type:", block)
    for chunk in chunks[1:]:
        chunk = "Type:" + chunk
        actions.append({
            "Type": parse_int(chunk, "Type"),
            "Target": parse_int(chunk, "Target"),
            "Value": parse_int(chunk, "Value"),
            "StatusId": parse_scalar(chunk, "StatusId"),
            "Stacks": parse_int(chunk, "Stacks", 1),
            "Duration": parse_int(chunk, "Duration", -1),
            "ScaleWithAttack": parse_int(chunk, "ScaleWithAttack"),
            "ScaleWithDefense": parse_int(chunk, "ScaleWithDefense"),
            "AttackScalePercent": parse_int(chunk, "AttackScalePercent", 100),
            "DefenseScalePercent": parse_int(chunk, "DefenseScalePercent", 100),
            "Condition": parse_int(chunk, "Condition"),
            "Reach": parse_int(chunk, "Reach", 1),
            "BackRowPowerPercent": parse_int(chunk, "BackRowPowerPercent", 100),
        })
    return actions


def format_action(a):
    parts = []
    cond = CONDITION.get(a["Condition"], "")
    if cond:
        parts.append(cond)

    reach = REACH.get(a["Reach"], "")
    t = ACTION.get(a["Type"], str(a["Type"]))
    tgt = TARGET.get(a["Target"], "")

    if a["Type"] == 0:
        dmg = f"ATK×{a['AttackScalePercent']//100}+{a['Value']}" if a["ScaleWithAttack"] else str(a["Value"])
        slot = f"【{reach}】" if reach else ""
        extra = ""
        if a["BackRowPowerPercent"] != 100:
            extra = f"，后排威力{a['BackRowPowerPercent']}%"
        parts.append(f"{slot}对{tgt}造成{dmg}伤害{extra}".replace("对默认敌人", "对敌人"))
    elif a["Type"] == 1:
        if a["ScaleWithDefense"]:
            val = f"DEF×{a['DefenseScalePercent']//100}"
        else:
            val = f"{a['Value']}点"
        parts.append(f"获得{val}护甲")
    elif a["Type"] == 2:
        parts.append(f"恢复{a['Value']}HP")
    elif a["Type"] == 3:
        sid = STATUS_CN.get(a["StatusId"], a["StatusId"] or "状态")
        dur = "永久" if a["Duration"] == -1 else f"{a['Duration']}回合"
        parts.append(f"对{tgt}施加{sid}×{a['Stacks']}（{dur}）")
    elif a["Type"] == 11:
        parts.append("随机使一名敌人本回合无法出牌")
    elif a["Type"] == 12:
        parts.append(f"玩家下回合能量回复-{a['Value']}")
    elif a["Type"] == 13:
        parts.append("【应对攻击】将伤害×2转嫁给随机队友")
    elif a["Type"] == 9:
        parts.append(f"获得{a['Value']}%减伤")
    else:
        parts.append(f"{t}→{tgt}({a['Value']})")
    return " ".join(p for p in parts if p)


def format_card_effect(card):
    return " → ".join(format_action(a) for a in card["Actions"])


def load_cards(card_dir, card_guids):
    cards = {}
    for fname in os.listdir(card_dir):
        if not fname.startswith("Card_") or not fname.endswith(".asset"):
            continue
        text = read_text(os.path.join(card_dir, fname))
        if "CardDefinitionSO" not in text:
            continue
        owner = parse_scalar(text, "OwnerCharacterId")
        card = {
            "Id": parse_scalar(text, "CardId"),
            "Name": parse_scalar(text, "DisplayName"),
            "Owner": owner,
            "Cost": parse_int(text, "Cost"),
            "Type": parse_int(text, "CardType"),
            "Rarity": parse_int(text, "Rarity", 0),
            "Keywords": [k for k in parse_list(text, "Keywords") if k],
            "Actions": parse_actions(text),
            "File": fname,
        }
        meta_path = os.path.join(card_dir, fname + ".meta")
        if os.path.exists(meta_path):
            gm = re.search(r"^guid:\s*(\w+)", read_text(meta_path), re.M)
            if gm:
                cards[gm.group(1)] = card
        cards[card["Id"]] = card
    return cards


def format_traits(traits):
    if not traits:
        return "无"
    return "；".join(TRAIT_CN.get(t, t) for t in traits)


def load_characters(char_dir):
    chars = []
    guid_to_char = {}
    for fname in os.listdir(char_dir):
        if not fname.startswith("Character_") or not fname.endswith(".asset"):
            continue
        path = os.path.join(char_dir, fname)
        text = read_text(path)
        if "CharacterDefinitionSO" not in text:
            continue
        if parse_int(text, "Team", 0) != 1:
            continue
        cid = parse_scalar(text, "CharacterId")
        deck_guids = parse_guids(parse_list(text, "Deck"))
        pool_guids = parse_guids(parse_list(text, "SkillPool"))
        char = {
            "Asset": fname,
            "Id": cid,
            "Name": parse_scalar(text, "DisplayName"),
            "Slot": parse_int(text, "Slot", 1),
            "MaxHp": parse_int(text, "MaxHp"),
            "Atk": parse_int(text, "BaseAttack"),
            "Def": parse_int(text, "BaseDefense"),
            "Spd": parse_int(text, "Speed"),
            "DeckGuids": deck_guids,
            "PoolGuids": pool_guids,
            "DeckSize": parse_int(text, "EnemyRandomDeckSize", 8),
            "PickMin": parse_int(text, "EnemySkillPickMin", 2),
            "PickMax": parse_int(text, "EnemySkillPickMax", 4),
            "Traits": [t for t in parse_list(text, "Traits") if t],
            "IsBoss": cid in BOSS_IDS,
        }
        chars.append(char)
        meta_path = path + ".meta"
        if os.path.exists(meta_path):
            gm = re.search(r"^guid:\s*(\w+)", read_text(meta_path), re.M)
            if gm:
                guid_to_char[gm.group(1)] = char
    return chars, guid_to_char


def is_player_asset(fname):
    return any(fname.startswith(p) for p in PLAYER_ASSET_PREFIXES)


def load_encounters(setup_dir, guid_to_char, char_guids):
    encounters = []
    for fname in sorted(os.listdir(setup_dir)):
        if not fname.startswith("BattleSetup") or not fname.endswith(".asset"):
            continue
        text = read_text(os.path.join(setup_dir, fname))
        if "BattleSetupSO" not in text:
            continue
        guids = parse_guids(parse_list(text, "Combatants"))
        summon_guids = parse_guids(parse_list(text, "SummonTemplates"))
        enemies = []
        for g in guids:
            ch = guid_to_char.get(g)
            if ch:
                enemies.append(ch["Name"])
                continue
            aname = char_guids.get(g, "")
            if aname and not is_player_asset(aname):
                enemies.append(aname.replace("Character_", "").replace("_", " "))
        summons = []
        for g in summon_guids:
            ch = guid_to_char.get(g)
            if ch:
                summons.append(ch["Name"])
        energy = parse_int(text, "EnemyTurnEnergyBudget")
        draw = parse_int(text, "EnemyCardsDrawnPerTurn")
        regen = parse_int(text, "TurnStartEnergyRegen", 4)
        encounters.append({
            "Name": fname.replace(".asset", ""),
            "Enemies": enemies,
            "Summons": summons,
            "Energy": energy if energy > 0 else regen,
            "Draw": draw if draw > 0 else parse_int(text, "CardsDrawnPerTurn", 5),
            "SkipScale": parse_int(text, "SkipFloorScaling") == 1,
        })
    return encounters


def monster_sort_key(c):
    order = ["char_goblin", "char_slime", "char_skeleton", "char_skeleton_elite", "char_wraith", "char_wraith_elite"]
    if c["Id"] in order:
        return (0, order.index(c["Id"]))
    return (1, c["Id"])


def write_cell(ws, row, col, value, bold=False, wrap=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.alignment = Alignment(wrap_text=wrap, vertical="top")
    if bold:
        cell.font = Font(bold=True)
    return cell


def write_boss_cards(ws, row, boss, cards_by_guid, energy_label=None, draw_label=None):
    headers = ["Boss卡牌", "卡牌ID", "卡牌名称", "费用", "类型", "站位/关键词", "卡牌稀有度", "卡牌数量", "效果描述", "备注"]
    for i, h in enumerate(headers, 1):
        write_cell(ws, row, i, h, bold=True)
    row += 1

    counter = Counter()
    for g in boss["DeckGuids"]:
        c = cards_by_guid.get(g)
        if c:
            counter[c["Id"]] += 1
    seen = set()
    for g in boss["DeckGuids"]:
        c = cards_by_guid.get(g)
        if not c or c["Id"] in seen:
            continue
        seen.add(c["Id"])
        kws = "、".join(KW_CN.get(k, k) for k in c["Keywords"]) or "Boss卡"
        cnt = counter[c["Id"]]
        write_cell(ws, row, 1, boss["Name"])
        write_cell(ws, row, 2, c["Id"])
        write_cell(ws, row, 3, c["Name"])
        write_cell(ws, row, 4, c["Cost"])
        write_cell(ws, row, 5, CARD_TYPE.get(c["Type"], c["Type"]))
        write_cell(ws, row, 6, kws)
        write_cell(ws, row, 7, RARITY.get(c["Rarity"], "白"))
        write_cell(ws, row, 8, cnt if cnt > 1 else (cnt if boss["Id"] != "char_explosive_skull" else "-"))
        write_cell(ws, row, 9, format_card_effect(c))
        note = ""
        if boss["Id"] == "char_explosive_skull":
            note = "仅易爆骷髅头在场时，必然在回合开始时加入手牌，不占抽牌数"
        write_cell(ws, row, 10, note)
        row += 1
    return row


def write_boss_sheet(wb, bosses, boss_encounters, cards_by_guid):
    if "Boss设计" in wb.sheetnames:
        ws = wb["Boss设计"]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet("Boss设计")

    boss_by_id = {b["Id"]: b for b in bosses}
    row = 1
    write_cell(ws, row, 1, "关于能量：Boss有不同的能量上限，但只要特性没特别说明，都是每回合直接回满能量；抽牌数量同理。")
    row += 2

    def write_boss_stats_section(encounter_name, boss_ids, shared_energy=None, shared_draw=None):
        nonlocal row
        enc = next((e for e in boss_encounters if e["Name"] == encounter_name), None)
        energy = enc["Energy"] if enc else shared_energy
        draw = enc["Draw"] if enc else shared_draw

        headers = ["Boss基本数值", "角色ID", "HP", "ATK", "DEF", "SPD", "总能量上限", "每回合抽牌数", "Boss特性", "备注"]
        for i, h in enumerate(headers, 1):
            write_cell(ws, row, i, h, bold=True)
        row += 1

        for i, bid in enumerate(boss_ids):
            b = boss_by_id.get(bid)
            if not b:
                continue
            e_label = energy if i == 0 or bid != "char_explosive_skull" else "共享骷髅王"
            d_label = draw if i == 0 or bid != "char_explosive_skull" else "共享骷髅王"
            if bid == "char_explosive_skull":
                e_label = d_label = "共享骷髅王"
            write_cell(ws, row, 1, b["Name"])
            write_cell(ws, row, 2, b["Id"])
            write_cell(ws, row, 3, b["MaxHp"])
            write_cell(ws, row, 4, b["Atk"])
            write_cell(ws, row, 5, b["Def"])
            write_cell(ws, row, 6, b["Spd"])
            write_cell(ws, row, 7, e_label)
            write_cell(ws, row, 8, d_label)
            write_cell(ws, row, 9, format_traits(b["Traits"]))
            note = "召唤物" if bid == "char_explosive_skull" else ""
            write_cell(ws, row, 10, note)
            row += 1

        row += 1
        for bid in boss_ids:
            b = boss_by_id.get(bid)
            if b and b["DeckGuids"]:
                row = write_boss_cards(ws, row, b, cards_by_guid)
        row += 1

    write_boss_stats_section(
        "BattleSetup_Encounter_SkeletonKingBoss",
        ["char_skeleton_king", "char_explosive_skull"],
    )
    write_boss_stats_section(
        "BattleSetup_Encounter_GhostQueenBoss",
        ["char_ghost_queen"],
    )

    enc_headers = ["Boss遭遇", "遭遇ID", "Boss阵容", "召唤模板", "能量", "抽牌", "层数缩放"]
    for i, h in enumerate(enc_headers, 1):
        write_cell(ws, row, i, h, bold=True)
    row += 1
    for e in boss_encounters:
        write_cell(ws, row, 1, e["Name"].replace("BattleSetup_Encounter_", "").replace("_", " "))
        write_cell(ws, row, 2, e["Name"])
        write_cell(ws, row, 3, " + ".join(e["Enemies"]))
        write_cell(ws, row, 4, " + ".join(e["Summons"]) if e["Summons"] else "无")
        write_cell(ws, row, 5, e["Energy"])
        write_cell(ws, row, 6, e["Draw"])
        write_cell(ws, row, 7, "否" if e["SkipScale"] else "是")
        row += 1

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["I"].width = 52
    ws.column_dimensions["J"].width = 36


def main():
    char_dir = os.path.join(PROJECT, "Assets/_Project/Data/Characters")
    card_dir = os.path.join(PROJECT, "Assets/_Project/Data/Cards")
    setup_dir = os.path.join(PROJECT, "Assets/_Project/Data/Setups")

    char_guids = build_guid_map(char_dir)
    cards_by_guid = load_cards(card_dir, char_guids)

    all_chars, guid_to_char = load_characters(char_dir)
    monsters = [c for c in all_chars if not c["IsBoss"]]
    bosses = [c for c in all_chars if c["IsBoss"]]
    monsters.sort(key=monster_sort_key)
    bosses.sort(key=lambda b: ["char_skeleton_king", "char_explosive_skull", "char_ghost_queen"].index(b["Id"])
                if b["Id"] in BOSS_IDS else 99)

    encounters = load_encounters(setup_dir, guid_to_char, char_guids)
    boss_encounters = [e for e in encounters if "Boss" in e["Name"]]
    normal_encounters = [e for e in encounters if "Boss" not in e["Name"]]

    shutil.copy2(REF_XLSX, OUT_XLSX)
    wb = load_workbook(OUT_XLSX)
    if "小怪设计" in wb.sheetnames:
        ws = wb["小怪设计"]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet("小怪设计")

    row = 1
    write_cell(ws, row, 1, "关于能量：小怪每回合最多 4 点能量、抽 5 张牌（Demo 与玩家相同）。")
    row += 1
    write_cell(ws, row, 1, "关于牌组：每只小怪的 SkillPool 按配置逐张加入（默认每种 1 张，重复条目即多张）；全场小怪牌汇入同一抽牌堆并洗牌。Boss 使用固定 Deck。")
    row += 1
    write_cell(ws, row, 1, "关于缩放：远征战斗按层数缩放 HP+15%/层、ATK+10%/层、DEF+5%/层（Boss 遭遇 SkipFloorScaling）。")
    row += 2

    headers = ["小怪基本数值", "角色ID", "HP", "ATK", "DEF", "SPD", "默认站位", "每回合能量", "每回合抽牌", "牌库规则", "特性", "备注"]
    for i, h in enumerate(headers, 1):
        write_cell(ws, row, i, h, bold=True)
    row += 1

    for m in monsters:
        if m["PoolGuids"]:
            deck_rule = f"技能池：{len(m['PoolGuids'])} 项（默认每项 1 张，可重复配置数量）"
        elif m["DeckGuids"]:
            deck_rule = f"固定牌库：{len(m['DeckGuids'])}张"
        else:
            deck_rule = "无"
        write_cell(ws, row, 1, m["Name"])
        write_cell(ws, row, 2, m["Id"])
        write_cell(ws, row, 3, m["MaxHp"])
        write_cell(ws, row, 4, m["Atk"])
        write_cell(ws, row, 5, m["Def"])
        write_cell(ws, row, 6, m["Spd"])
        write_cell(ws, row, 7, SLOT.get(m["Slot"], m["Slot"]))
        write_cell(ws, row, 8, 4)
        write_cell(ws, row, 9, 5)
        write_cell(ws, row, 10, deck_rule)
        write_cell(ws, row, 11, format_traits(m["Traits"]))
        write_cell(ws, row, 12, "")
        row += 1

    row += 1
    card_headers = ["小怪卡牌", "卡牌ID", "卡牌名称", "费用", "类型", "站位/关键词", "卡牌稀有度", "技能池/固定", "效果描述", "备注"]
    for i, h in enumerate(card_headers, 1):
        write_cell(ws, row, i, h, bold=True)
    row += 1

    for m in monsters:
        pool_cards = []
        if m["PoolGuids"]:
            for g in m["PoolGuids"]:
                c = cards_by_guid.get(g)
                if c:
                    pool_cards.append(c)
            pool_cards.sort(key=lambda x: (x["Cost"], x["Id"]))
            for c in pool_cards:
                kws = "、".join(KW_CN.get(k, k) for k in c["Keywords"])
                write_cell(ws, row, 1, m["Name"])
                write_cell(ws, row, 2, c["Id"])
                write_cell(ws, row, 3, c["Name"])
                write_cell(ws, row, 4, c["Cost"])
                write_cell(ws, row, 5, CARD_TYPE.get(c["Type"], c["Type"]))
                write_cell(ws, row, 6, kws)
                write_cell(ws, row, 7, RARITY.get(c["Rarity"], "白"))
                write_cell(ws, row, 8, "技能池")
                write_cell(ws, row, 9, format_card_effect(c))
                write_cell(ws, row, 10, c["File"])
                row += 1
        if m["DeckGuids"]:
            counter = Counter()
            for g in m["DeckGuids"]:
                c = cards_by_guid.get(g)
                if c:
                    counter[c["Id"]] += 1
            seen = set()
            for g in m["DeckGuids"]:
                c = cards_by_guid.get(g)
                if not c or c["Id"] in seen:
                    continue
                seen.add(c["Id"])
                kws = "、".join(KW_CN.get(k, k) for k in c["Keywords"])
                cnt = counter[c["Id"]]
                write_cell(ws, row, 1, m["Name"])
                write_cell(ws, row, 2, c["Id"])
                write_cell(ws, row, 3, c["Name"])
                write_cell(ws, row, 4, c["Cost"])
                write_cell(ws, row, 5, CARD_TYPE.get(c["Type"], c["Type"]))
                write_cell(ws, row, 6, kws)
                write_cell(ws, row, 7, RARITY.get(c["Rarity"], "白"))
                write_cell(ws, row, 8, f"固定×{cnt}")
                write_cell(ws, row, 9, format_card_effect(c))
                write_cell(ws, row, 10, c["File"])
                row += 1

    row += 1
    enc_headers = ["遭遇组合", "遭遇ID", "敌方阵容（3人）", "每回合能量", "每回合抽牌", "层数缩放", "备注"]
    for i, h in enumerate(enc_headers, 1):
        write_cell(ws, row, i, h, bold=True)
    row += 1

    name_map = {
        "Goblin": "哥布林", "Slime": "史莱姆", "Skeleton": "骷髅兵", "Skeleton_Elite": "骷髅精英",
        "Wraith": "幽灵", "Wraith_Elite": "幽灵精英",
    }
    for e in normal_encounters:
        lineup = " + ".join(e["Enemies"])
        note = "Demo 远征池" if e["Name"] in ("BattleSetup_Demo", "BattleSetup_Encounter_SlimeMix", "BattleSetup_Encounter_WraithPack") else ""
        write_cell(ws, row, 1, e["Name"].replace("BattleSetup", "").replace("_", " ").strip())
        write_cell(ws, row, 2, e["Name"])
        write_cell(ws, row, 3, lineup)
        write_cell(ws, row, 4, e["Energy"])
        write_cell(ws, row, 5, e["Draw"])
        write_cell(ws, row, 6, "否" if e["SkipScale"] else "是")
        write_cell(ws, row, 7, note)
        row += 1

    row += 1
    write_cell(ws, row, 1, "技能池组合说明（每怪 2~4 种 × 8 张牌库，以下为各怪技能池全部卡牌）", bold=True)
    row += 1
    for m in monsters:
        if not m["PoolGuids"]:
            continue
        names = []
        for g in m["PoolGuids"]:
            c = cards_by_guid.get(g)
            if c:
                names.append(c["Name"])
        write_cell(ws, row, 1, m["Name"])
        write_cell(ws, row, 2, f"共{len(names)}种：{' / '.join(names)}")
        write_cell(ws, row, 3, f"C({len(names)},2)+C({len(names)},3)+C({len(names)},4) 种组合（实际每局随机 2~4 种）")
        row += 1

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 48
    ws.column_dimensions["J"].width = 28
    ws.column_dimensions["K"].width = 36
    ws.column_dimensions["L"].width = 18

    write_boss_sheet(wb, bosses, boss_encounters, cards_by_guid)

    wb.save(OUT_XLSX)
    import sys
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
    print(f"Saved: {OUT_XLSX}")
    print(f"Monsters: {len(monsters)}, Bosses: {len(bosses)}, rows: {row}")


if __name__ == "__main__":
    main()
