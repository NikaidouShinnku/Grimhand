# -*- coding: utf-8 -*-
import openpyxl
import pathlib
import re

wb = openpyxl.load_workbook(
    r"c:\Users\Kelthuzad\Desktop\Grimhand实际内容总览表v0.93.xlsx", data_only=True
)
ws = wb["卡牌"]
cards = []
for r in range(2, ws.max_row + 1):
    name = ws.cell(r, 2).value
    maxu = ws.cell(r, 8).value
    effect = ws.cell(r, 9).value
    xp = ws.cell(r, 10).value
    if not name:
        continue
    if not maxu or str(maxu).strip() in ("-", "—", "无", ""):
        continue
    try:
        maxi = int(str(maxu).strip())
    except ValueError:
        continue
    xp_n = None
    if xp:
        m = re.search(r"(\d+)", str(xp))
        if m:
            xp_n = int(m.group(1))
    cards.append((str(name), maxi, str(effect or ""), xp_n))

text = pathlib.Path(r"Assets/_Project/Scripts/Core/CardUpgradeCatalog.cs").read_text(
    encoding="utf-8"
)
pat = re.compile(r'\["([^"]+)"\]\s*=\s*new\(\)\s*\{([^}]+)\}')
cat = {}
for m in pat.finditer(text):
    name = m.group(1)
    body = m.group(2)

    def get(field, default=0):
        mm = re.search(field + r"\s*=\s*(-?\d+)", body)
        return int(mm.group(1)) if mm else default

    cat[name] = {
        "Max": get("MaxUpgrades"),
        "Dmg": get("DamagePerLevel"),
        "Blk": get("BlockPerLevel"),
        "Heal": get("HealPerLevel"),
        "Cost": get("CostReductionPerLevel"),
        "Poi": get("PoisonStacksPerLevel"),
        "Slow": get("SlowStacksPerLevel"),
        "Draw": get("DrawPerLevel"),
        "DR": get("DamageReductionPerLevel"),
        "Mit": get("RespondMitigationPerLevel"),
        "Ref": get("ReflectPercentPerLevel"),
        "Xp": get("XpCostPerLevel"),
    }

print("=== MISSING FROM CATALOG ===")
for name, maxi, effect, xp in cards:
    if name not in cat:
        print(f"{name}|max={maxi}|xp={xp}|{effect}")

print("\n=== MAX/XP MISMATCH ===")
for name, maxi, effect, xp in cards:
    if name not in cat:
        continue
    c = cat[name]
    issues = []
    if c["Max"] != maxi:
        issues.append(f"Max {c['Max']}!={maxi}")
    if xp is not None and c["Xp"] != xp:
        issues.append(f"Xp {c['Xp']}!={xp}")
    if issues:
        print(f"{name}|{effect}|{'; '.join(issues)}")

print("\n=== SIMPLE EFFECT HEURISTICS ===")
for name, maxi, effect, xp in cards:
    if name not in cat:
        continue
    c = cat[name]
    e = effect.strip()
    issues = []
    if e == "+1伤害" and c["Dmg"] != 1:
        issues.append(f"Dmg={c['Dmg']}")
    if e == "+2伤害" and c["Dmg"] != 2:
        issues.append(f"Dmg={c['Dmg']}")
    if e == "+3伤害" and c["Dmg"] != 3:
        issues.append(f"Dmg={c['Dmg']}")
    if e == "+4伤害" and c["Dmg"] != 4:
        issues.append(f"Dmg={c['Dmg']}")
    if e == "+5伤害" and c["Dmg"] != 5:
        issues.append(f"Dmg={c['Dmg']}")
    if e == "+10伤害" and c["Dmg"] != 10:
        issues.append(f"Dmg={c['Dmg']}")
    if e == "+1护甲" and c["Blk"] != 1:
        issues.append(f"Blk={c['Blk']}")
    if e == "+2护甲" and c["Blk"] != 2:
        issues.append(f"Blk={c['Blk']}")
    if e in ("+1HP", "+1回复HP", "+1恢复HP") and c["Heal"] != 1:
        issues.append(f"Heal={c['Heal']}")
    if e == "+5恢复HP" and c["Heal"] != 5:
        issues.append(f"Heal={c['Heal']}")
    if e == "+1层中毒" and c["Poi"] != 1:
        issues.append(f"Poi={c['Poi']}")
    if e == "+3层中毒" and c["Poi"] != 3:
        issues.append(f"Poi={c['Poi']}")
    if e == "+1抽牌数" and c["Draw"] != 1:
        issues.append(f"Draw={c['Draw']}")
    if e == "费用-1" and c["Cost"] != 1:
        issues.append(f"Cost={c['Cost']}")
    if e == "+5%减伤，+10%反弹伤害":
        if c["Mit"] != 5 or c["Ref"] != 10:
            issues.append(f"Mit={c['Mit']} Ref={c['Ref']}")
    if e == "+5%减伤" and c["Mit"] != 5 and c["DR"] != 5:
        issues.append(f"减伤 Mit={c['Mit']} DR={c['DR']}")
    if e == "+2%减伤" and c["DR"] != 2 and c["Mit"] != 2:
        issues.append(f"减伤 Mit={c['Mit']} DR={c['DR']}")
    if issues:
        print(f"{name}|{e}|{'; '.join(issues)}")

# relics growth
print("\n=== RELIC GROWTH (excel) ===")
ws = wb["遗物"]
for r in range(3, ws.max_row + 1):
    rid = ws.cell(r, 1).value
    name = ws.cell(r, 2).value
    growth = ws.cell(r, 7).value
    if rid:
        print(f"{rid}|{name}|{growth}")
