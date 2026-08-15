# -*- coding: utf-8 -*-
import openpyxl
ws = openpyxl.load_workbook(r"c:\Users\Kelthuzad\Desktop\Grimhand实际内容总览表v0.93.xlsx", data_only=True)["卡牌"]
for r in range(2, ws.max_row+1):
    name = ws.cell(r,2).value
    maxu = ws.cell(r,8).value
    effect = ws.cell(r,9).value
    xp = ws.cell(r,10).value
    if not name: continue
    if not maxu or str(maxu).strip() in ("-","—","无",""): continue
    print(f"{name}\t{maxu}\t{effect}\t{xp}")
